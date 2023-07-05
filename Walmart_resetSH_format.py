import streamlit as st
from openpyxl.styles import numbers, PatternFill, Border, Side, Font, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import datetime
import pandas as pd
import openpyxl
import tempfile
from datetime import datetime

def format_WALMART_schedule(workbook):
    st.write("YAY YOU CALLED ME WALMART")
    # Select the Reset Dates sheet
    ws = workbook['WALMART_RESET']

    # Insert a new column "Chain_Name" at the beginning and fill it with "WALMART"
    ws.insert_cols(1)
    ws.cell(row=1, column=1, value="Chain_Name")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.value = "WALMART"

    # Insert a new column "Store_Name" at the third position and fill it with "WALMART"
    ws.insert_cols(3)
    ws.cell(row=1, column=3, value="Store_Name")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.value = "WALMART"

    # Move column "ADDRESS" (previously column V) to the fifth position
    ws.move_range("V1:V{}".format(ws.max_row), cols=5)

    # Move column "State" (previously column U) to the sixth position
    ws.move_range("U1:U{}".format(ws.max_row), cols=6)

    # Move column "E" to the ninth position
    ws.move_range("E1:E{}".format(ws.max_row), cols=9)

    # Delete the current column F
    ws.delete_cols(6)

    # Insert a column for "PHONE"
    ws.insert_cols(4)
    # Insert a column for "PHONE"
    ws.insert_cols(5)

    # Move column "CITY" to the FIFTH position
    #ws.move_range("J1:J{}".format(ws.max_row), cols=5)
    # Copy values from column J to column E
    for cell in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=10, max_col=10):
        for c in cell:
            ws.cell(row=c.row, column=5, value=c.value)

    # Insert a column for "ADDRESS"
    ws.insert_cols(6)

    # Copy values from column J to column j
    for cell in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
        for c in cell:
            ws.cell(row=c.row, column=6, value=c.value)

    # Insert a column for "STATE"
    ws.insert_cols(7)

    # Copy values from column J to column G
    for cell in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=11, max_col=11):
        for c in cell:
            ws.cell(row=c.row, column=7, value=c.value)
    
    # Insert a column for "STATE"
    ws.insert_cols(8)

    # Insert a column for "TEAM LEAD"
    ws.insert_cols(9)

    # Copy values from column J to column I
    for cell in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
        for c in cell:
            ws.cell(row=c.row, column=9, value=c.value)

    # Copy values from column S to column K
    for cell in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=19, max_col=19):
        for c in cell:
            ws.cell(row=c.row, column=11, value=c.value)

    # Delete unnecessary columns
    ws.delete_cols(12, 14)  # Delete columns L, M, and N
    ws.delete_cols(15, 17)  # Delete columns O, P, and Q



    # Check if the worksheet has any data
    if ws.dimensions is not None:
        # Iterate over column J and format the date
        for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.value = cell.value.strftime('%m/%d/%Y')

    # Convert the column index to letter
    column_letter = get_column_letter(10)
    ws.insert_cols(11)
    # Create a named style for the date format
    date_format = NamedStyle(name='date_format')
    date_format.number_format = 'mm/dd/yyyy'  # Set the desired date format

    # Apply the named style to the cells in column J
    for cell in ws[column_letter]:
        cell.style = date_format

   # Save the workbook to a temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False)
    workbook.save(temp_file.name)
    temp_file.close()

        # Read the modified spreadsheet into a DataFrame
    df = pd.read_excel(temp_file.name, sheet_name='WALMART_RESET', header=0)

    # Get the column index for the desired columns
    column_index_time = 11  # Column L

    # Insert a new column for the formatted time before column L
    df.insert(column_index_time, 'New_Time_Column', '')

    # Format the time values in the new column
    df['New_Time_Column'] = df['Time'].apply(lambda time_value: time_value.strftime('%I:%M %p') if pd.notnull(time_value) else '')

    # Delete column 10 at the end
    df = df.drop(df.columns[10], axis=1)

    
    # Delete column 10 at the end
    df = df.drop(df.columns[11], axis=1)

    # Save the updated DataFrame back to the Excel file using xlsxwriter engine
    new_filename = 'formatted_WALMART_schedule.xlsx'
    writer = pd.ExcelWriter(new_filename, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='WALMART_RESET')
    writer.save()

    #st.write(df)

    # Check the current number of columns
    num_columns = len(df.columns)
    #st.write(df.columns)

    # Determine the number of additional columns needed
    num_additional_columns = 13 - num_columns

    # Add additional columns with default values
    for i in range(num_additional_columns):
        df[f'Column_{num_columns + i + 1}'] = ''

    # Assign column names to the DataFrame
    df.columns = [
        'CHAIN_NAME',
        'STORE_NUMBER',
        'STORE_NAME',
        'PHONE',
        'CITY',
        'ADDRESS',
        'STATE',
        'COUNTY',
        'TEAM_LEAD',
        'RESET_DATE',
        'RESET_TIME',
        'STATUS',
        'NOTES'
    ]
   
    # Print the updated column names
    print(df.columns)
    # Check if 'Status' column exists
    if 'Status' in df.columns:
        # Assign default values or placeholders to 'Status' column
        df['Status'].fillna('Unknown', inplace=True)

    # Check if 'Notes' column exists
    if 'Notes' in df.columns:
        # Assign default values or placeholders to 'Notes' column
        df['Notes'].fillna('', inplace=True)

    
        # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        # Specify the column you want to format (e.g., column A)
        column_letter = 'J'

    # Iterate over the cells in the column and set the number format
    for cell in ws[column_letter]:
        cell.number_format = 'mm/dd/yyyy'

    # Convert the DataFrame back to an Excel file
    new_filename = 'formatted_WALMART_schedule.xlsx'
    df.to_excel(new_filename, index=False, sheet_name='WALMART_RESET')

    #st.write(df.columns)

    # Return the modified workbook
    return openpyxl.load_workbook(new_filename)


















