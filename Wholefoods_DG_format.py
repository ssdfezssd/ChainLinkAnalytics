import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.styles import numbers
from openpyxl import load_workbook





# Reformat Excel spreadsheet for SAFEWAY
#====================================================================================================================
def format_WHOLEFOODS_DistroGrid(workbook):

    st.write("YAY YOU CALLED ME WHOLE FOODS dg CODE")


    # Select the distrogrid  sheet
    ws = workbook['MASTER CORE LIST_NC']

    # Insert a new column "Chain_Name" at the beginning and fill it with "WALMART"
    ws.insert_cols(1)
    ws.cell(row=1, column=1, value="STORE_NAME")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.value = "WHOLE FOODS"

    # Delete columns 2(B)
    ws.delete_cols(2)

    # Delete columns 2(B)
    ws.delete_cols(2)

    # Delete columns 2(B)
    ws.delete_cols(2)

     # Delete columns 2(B)
    ws.delete_cols(2)

     # Delete columns 2(B)
    ws.delete_cols(3)



    ## Remove the filter from the sheet
    ws.auto_filter.ref = None

    #    # Iterate over the rows starting from row 2 and remove all values 
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.value = None

    source_column = 'G'
    destination_column = 'D'

    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            destination_cell = ws.cell(row=cell.row, column=4)
            destination_cell.value = cell.value
            cell.value = None


     # Delete columns 2(B)
    ws.delete_cols(5)

    # Insert a column for "PHONE"
    ws.insert_cols(4)

     # Move column "G" to the fourth position
    #ws.move_range("G1:G{}".format(ws.max_row), cols=4)

    #   # Iterate over the rows starting from row 2 and remove all values 
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.value = None

    #   # Iterate over the rows starting from row 2 and remove all values 
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
        for cell in row:
            cell.value = None

    #   # Iterate over the rows starting from row 2 and remove all values 
    for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):
        for cell in row:
            cell.value = None


    # Delete columns 2(B)
    ws.delete_cols(12)

    # Assuming the worksheet is stored in the variable 'ws'
    column_c = ws['C']

    # Convert the values in column C to numeric format
    for cell in column_c:
        value = cell.value
        if value is not None:
            try:
                cell.value = float(value)
            except ValueError:
                pass

    # Set the number format for column C to whole number (no decimal places)
    for cell in column_c:
        if cell.value is not None:
            cell.number_format = '0'


    for cell in ws['H2:H{}'.format(ws.max_row)]:
        cell[0].value = 1



    # Rename Columns as required to meet objective for uploading to Snowflake
    ws.cell(row=1, column=1, value='STORE_Name')
    ws.cell(row=1, column=2, value='STORE_Number')
    ws.cell(row=1, column=3, value='UPC')
    ws.cell(row=1, column=4, value='SKU')
    ws.cell(row=1, column=5, value='PRODUCT_NAME')
    ws.cell(row=1, column=6, value='MANUFACTURER')
    ws.cell(row=1, column=7, value='SEGMENT')
    ws.cell(row=1, column=8, value='YES_NO')
    ws.cell(row=1, column=9, value='ACTIVATION_STATUS')
    ws.cell(row=1, column=10, value='COUNTY')
    ws.cell(row=1, column=11, value='CHAIN_NAME')

    # Insert a new column "Chain_Name" at the beginning and fill it with "WALMART"
    #ws.insert_cols(1)
    ws.cell(row=1, column=11, value="CHAIN_NAME")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=11, max_col=11):
        for cell in row:
            cell.value = "WHOLE FOODS"
            
    # Iterate through the rows and update the 6th column (index 5) to uppercase
    for row in ws.iter_rows(min_row=2):  # Start from the second row to skip headers
        cell = row[5]  # 6th column (index 5)
        cell.value = cell.value.upper()
 


    return workbook
