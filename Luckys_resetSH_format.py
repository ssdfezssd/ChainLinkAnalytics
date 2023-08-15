import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, NamedStyle, numbers    
from openpyxl.utils import get_column_letter
import datetime




#=======================================================================================================================
# Function to adjust the width of the columns in the sheet
#======================================================================================================================

def adjust_column_widths(worksheet):
    # Iterate over all columns
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        # Find the maximum length of data in each column
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass

        # Set the column width based on the maximum length
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

#=======================================================================================================================
# End of Function to adjust the width of the columns in the sheet
#======================================================================================================================


#===================================================================================================================
# Function to reformat Excel spreadsheet for Luckys
#====================================================================================================================
def format_LUCKYS_Schedule(workbook):

    st.write("Formatting workbook for LUCKY")

    # Select the Reset Dates sheet
    ws = workbook['LUCKY']

    # Remove filter from the worksheet
    ws.auto_filter.ref = None

    # Create a list to store merged cell ranges
    merged_cell_ranges = []

    # Collect merged cell ranges
    for merged_cell_range in ws.merged_cells.ranges:
        merged_cell_ranges.append(merged_cell_range)

    # Unmerge cells
    for merged_cell_range in merged_cell_ranges:
        ws.unmerge_cells(str(merged_cell_range))

    # Delete Rows 1 and 2
    rows_to_delete = [1, 2]
    for row_index in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_index)

    # Determine the number of rows with data
    sheet_name = "LUCKY"  # Update with the actual sheet name
    ws = workbook[sheet_name]
    max_row = ws.max_row
    for row in reversed(range(1, max_row + 1)):
        if ws.cell(row=row, column=3).value:
            max_row = row
            break

   # Delete extra rows beyond the data range
    if ws.max_row > max_row:
        ws.delete_rows(max_row + 1, ws.max_row - max_row)

    
    # Get the maximum row number in column C
    max_row_c = ws.max_row

    # Add "SAVE MART" to each cell in column C
    for row in range(2, max_row_c + 1):
        cell = ws.cell(row=row, column=1)
        cell.value = "SAVE MART"

    # Insert column C to hold Store Name "Luckys"
    ws.insert_cols(3)

    # Get the maximum row number in column C
    max_row_c = ws.max_row

    # Add "SAVE MART" to each cell in column C
    for row in range(2, max_row_c + 1):
        cell = ws.cell(row=row, column=3)
        cell.value = "LUCKYS"

    # Insert column D to hold Store Name "PHONE"
    ws.insert_cols(4)

    # Get the column letters for Column H (8) and Column D (4)
    column_h = get_column_letter(9)
    column_d = get_column_letter(4)

    # Iterate over each row and move the values from Column H to Column D
    for row in range(1, ws.max_row + 1):
        ws[f"{column_d}{row}"].value = ws[f"{column_h}{row}"].value

     # Insert column D to hold Store Name "PHONE"
    ws.insert_cols(4)

    # Insert column E to hold Store Name "CITY"
    ws.insert_cols(5)

    # Get the column letters for Column J and Column E
    column_j = get_column_letter(10)
    column_e = get_column_letter(5)

    # Iterate over each row and move the values from Column J to Column E
    for row in range(1, ws.max_row + 1):
        ws[f"{column_e}{row}"].value = ws[f"{column_j}{row}"].value

    # Get the column letter for column G
    column_g = get_column_letter(7)

    # Insert two new columns between F and G
    ws.insert_cols(7, amount=2)

    # Delete columns J, K, L, and M
    ws.delete_cols(10, amount=4)

    # Delete columns J, K
    ws.delete_cols(10, amount=2)

    # Delete columns N - X
    ws.delete_cols(14, amount=8)

    # Specify the column to update (e.g., column K)
    column = 'K'

    # Iterate over the cells in the column and update the values
    for cell in ws[column]:
        if isinstance(cell.value, str):  # Check if the cell value is a string
            value = cell.value.strip().upper()  # Remove leading/trailing spaces and convert to uppercase
            if value == '6AM':
                cell.value = '6:00'
            elif value == '5AM':
                cell.value = '5:00'
            elif value == '7AM':
                cell.value = '7:00'
            elif value == '8AM':
                cell.value = '8:00'

    # Apply the desired time format
    for cell in ws[column]:
        if isinstance(cell.value, (datetime.datetime, datetime.time)):  # Check if the cell value is a datetime or time object
            cell.number_format = 'h:mm AM/PM'

    ## Check if the worksheet has any data
    #if ws.dimensions is not None:
    #    # Iterate over column P to convert text to number and format as time
    #    for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):
    #        for cell in row:
    #            if cell.value:
    #                try:
    #                    cell.value = float(cell.value)
    #                    cell.number_format = 'h:mm AM/PM'  # Add AM/PM format
    #                except ValueError:
    #                    print(f"Failed to convert cell {cell.coordinate} to float.")
    

          # Apply gridlines to all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

            # Remove cell colors
            cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

    # Make the entire sheet have black-colored text with font size 11
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(color="00000000", size=11)
   
    # Rename Columns as required to meet objective for uploading to Snowflake
    ws.cell(row=1, column=1, value='CHAIN_NAME')
    ws.cell(row=1, column=2, value='STORE_NUMBER')
    ws.cell(row=1, column=3, value='STORE_NAME')
    ws.cell(row=1, column=4, value='PHONE')
    ws.cell(row=1, column=5, value='CITY')
    ws.cell(row=1, column=6, value='ADDRESS')
    ws.cell(row=1, column=7, value='STATE')
    ws.cell(row=1, column=8, value='COUNTY')
    ws.cell(row=1, column=9, value='TEAM_LEAD')
    ws.cell(row=1, column=10, value='RESET_DATE')
    ws.cell(row=1, column=11, value='RESET_TIME')
    ws.cell(row=1, column=12, value='STATUS')
    ws.cell(row=1, column=13, value='NOTES')


    # Specify the column you want to format (e.g., column A)
    column_letter = 'J'

    # Iterate over the cells in the column and set the number format
    for cell in ws[column_letter]:
        cell.number_format = 'mm/dd/yyyy'


    # Get the maximum row number in column A
    max_row = ws.max_row

    # Add "CA" to each cell in column G
    for row in range(2, max_row_c + 1):
        cell = ws.cell(row=row, column=7)
        cell.value = "CA"
    
         #Find the last row with data in column A
    last_row = ws.max_row
    for row in reversed(range(1, last_row + 1)):
        if ws.cell(row=row, column=1).value:
            last_row = row
            break

        # Fill in empty cells in column J with '01/01/1900' or replace '#N/A' with '01/01/1900'
    for row in range(1, last_row + 1):
        if ws.cell(row=row, column=10).value == '#N/A':
            ws.cell(row=row, column=10).value = '01/01/1900'
            ws.cell(row=row, column=10).number_format = 'mm/dd/yyyy'
        elif not ws.cell(row=row, column=10).value:
            ws.cell(row=row, column=10).value = '01/01/1900'
            ws.cell(row=row, column=10).number_format = 'mm/dd/yyyy'
    
    adjust_column_widths(ws)


    # Return the path to the saved file
    return workbook


#===================================================================================================================
# End Function to reformat Excel spreadsheet for Luckys
#====================================================================================================================