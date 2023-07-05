import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.styles import numbers
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill, Font, NamedStyle
from openpyxl.styles import Font





# Reformat Excel spreadsheet for SAFEWAY
#====================================================================================================================
def format_SPROUTS_DistroGrid(workbook):

    st.write("YAY YOU CALLED ME SPROUTS dg CODE")

    import datetime

       # Get the current time
    current_time = datetime.datetime.now()

        # Format the current time as a string
    time_string = current_time.strftime("%Y-%m-%d %H:%M:%S")

        # Write the current time
    st.write("Current time:", time_string)


    # Select the Reset Dates sheet
    ws = workbook['SPROUTS_dg']

    # Insert a new column "Chain_Name" at the beginning and fill it with "WALMART"
    ws.insert_cols(1)
    ws.cell(row=1, column=1, value="STORE_NAME")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.value = ""

    ## Remove the filter from the sheet
    ws.auto_filter.ref = None

    ## Delete columns 2(B)
    ws.delete_cols(2)

    ## Insert a column for "UPC"
    ws.insert_cols(3)

        # Get the maximum row number
    max_row = ws.max_row

    # Iterate over the rows starting from the second row
    for row in range(2, max_row + 1):
        # Get the values from column I and assign them to column C
        value = ws.cell(row=row, column=column_index_from_string('I')).value
        ws.cell(row=row, column=column_index_from_string('C')).value = value

        # Clear the data in column I
        ws.cell(row=row, column=column_index_from_string('I')).value = None

    # Assuming the worksheet is stored in the variable 'ws'
    column_c = ws['C']

    # Convert the values in column C to numeric format
    for cell in column_c:
        value = cell.value
        if value is not None:
            try:
                cell.value = float(value)
            except ValueError:
                pass

    # Set the number format for column C to whole number (no decimal places)
    for cell in column_c:
        if cell.value is not None:
            cell.number_format = '0'

    ## Delete columns 6
    ws.delete_cols(6)

    ## Delete columns 6
    ws.delete_cols(6)
    ## Delete columns 6
    ws.delete_cols(6)
    ## Delete columns 6
    ws.delete_cols(6)
    ## Delete columns 6
    ws.delete_cols(6)

    # Insert a column for "SKU"
    ws.insert_cols(4)

    # Insert a column for "SEGMENT"
    ws.insert_cols(7)



    # Get the maximum row number
    max_row = ws.max_row

    # Iterate over the rows starting from the first row
    for row in range(1, max_row + 1):
        # Get the values from columns E and F
        value_e = ws.cell(row=row, column=5).value
        value_f = ws.cell(row=row, column=6).value

        # Assign the values to columns F and E, respectively
        ws.cell(row=row, column=5).value = value_f
        ws.cell(row=row, column=6).value = value_e


      ## ## CONVERT ALL ROWS IN COLUMN H TO 1 IF THE VALUE IS ADD
    for cell in ws['H']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('ADD', '1')

          ## CONVERT ALL ROWS IN COLUMN H TO 1 IF THE VALUE IS KEEP
    for cell in ws['H']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('KEEP', '1')

    # Define the translation table to remove characters
    translation_table = str.maketrans("", "", "'\u2018\u2019\u201B")

    # Remove single quotes from column E
    for cell in ws['E']:
        if cell.value is not None:
            cell.value = str(cell.value).translate(translation_table)

            # Define the translation table to remove characters
    translation_table = str.maketrans("", "", "'\u2018\u2019\u201B")

    # Remove single quotes from column E
    for cell in ws['F']:
        if cell.value is not None:
            cell.value = str(cell.value).translate(translation_table)
    

       # Rename Columns as required to meet objective for uploading to Snowflake
    ws.cell(row=1, column=1, value='STORE_Name')
    ws.cell(row=1, column=2, value='STORE_Number')
    ws.cell(row=1, column=3, value='UPC')
    ws.cell(row=1, column=4, value='SKU')
    ws.cell(row=1, column=5, value='PRODUCT_NAME')
    ws.cell(row=1, column=6, value='MANUFACTURER')
    ws.cell(row=1, column=7, value='SEGMENT')
    ws.cell(row=1, column=8, value='YES_NO')
    ws.cell(row=1, column=9, value='ACTIVATION_STATUS')
    ws.cell(row=1, column=10, value='COUNTY')
    ws.cell(row=1, column=11, value='CHAIN_NAME')

    
    # Convert the worksheet data to a Pandas DataFrame
    data = ws.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns)

    # Remove rows with "DELETE" in column H
    df = df[df['YES_NO'] != 'DELETE'].reset_index(drop=True)

    st.write(df)

    # Write the modified DataFrame back to the worksheet
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    # Adjust the column widths if needed
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.coordinate in ws.merged_cells:
                continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
   
    
    
    
    ## Delete columns 2(B)
    #ws.delete_cols(2)

    # # Delete columns 2(B)
    #ws.delete_cols(2)

    # # Delete columns 2(B)
    #ws.delete_cols(3)



    ### Remove the filter from the sheet
    #ws.auto_filter.ref = None

    ##    # Iterate over the rows starting from row 2 and remove all values 
    #for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
    #    for cell in row:
    #        cell.value = None

    #source_column = 'G'
    #destination_column = 'D'

    #for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
    #    for cell in row:
    #        destination_cell = ws.cell(row=cell.row, column=4)
    #        destination_cell.value = cell.value
    #        cell.value = None


    # # Delete columns 2(B)
    #ws.delete_cols(5)

    ## Insert a column for "PHONE"
    #ws.insert_cols(4)

    # # Move column "G" to the fourth position
    ##ws.move_range("G1:G{}".format(ws.max_row), cols=4)

    ##   # Iterate over the rows starting from row 2 and remove all values 
    #for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
    #    for cell in row:
    #        cell.value = None

    ##   # Iterate over the rows starting from row 2 and remove all values 
    #for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
    #    for cell in row:
    #        cell.value = None

    ##   # Iterate over the rows starting from row 2 and remove all values 
    #for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):
    #    for cell in row:
    #        cell.value = None


    ## Delete columns 2(B)
    #ws.delete_cols(12)

    ## Assuming the worksheet is stored in the variable 'ws'
    #column_c = ws['C']

    ## Convert the values in column C to numeric format
    #for cell in column_c:
    #    value = cell.value
    #    if value is not None:
    #        try:
    #            cell.value = float(value)
    #        except ValueError:
    #            pass

    ## Set the number format for column C to whole number (no decimal places)
    #for cell in column_c:
    #    if cell.value is not None:
    #        cell.number_format = '0'


    #for cell in ws['H2:H{}'.format(ws.max_row)]:
    #    cell[0].value = 1


    
          # Apply gridlines to all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

            # Remove cell colors
            cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

    # Make the entire sheet have black-colored text with font size 11
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(color="00000000", size=11)

    # Insert a new column "Chain_Name" at the beginning and fill it with "WALMART"
    #ws.insert_cols(1)
    ws.cell(row=1, column=11, value="CHAIN_NAME")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=11, max_col=11):
        for cell in row:
            cell.value = "SPROUTS"
 


    return workbook

