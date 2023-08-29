from ctypes.wintypes import SIZE
import streamlit as st
import snowflake.connector
import snowflake
import Distro_Grid_Snowflake_Uploader
import datetime
import pandas as pd
from PIL import Image
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import openpyxl.utils.datetime as xl_datetime
import numpy as np
from io import BytesIO
from openpyxl import Workbook
import base64
from datetime import time
from Safeway_resetSH_format import format_SAFEWAY_Schedule
from Raleys_resetSH_format import format_RALEYS_Schedule
from Walmart_resetSH_format import format_WALMART_schedule
from FoodMaxx_resetSH_format import format_FOODMAXX_schedule
from Luckys_resetSH_format import format_LUCKYS_Schedule
from Savemart_resetSH_format import format_Savemart_Schedule
from Sprouts_resetSH_format import format_SPROUTS_Schedule
from Smart_Final_resetSH_format import format_SMARTFINAL_Schedule
from Reset_Schedule_to_Snowflake_Uploader import upload_reset_SCH_SAFEWAY_data
from Reset_Schedule_to_Snowflake_Uploader import upload_reset_SCH_LUCKY_data
from Reset_Schedule_to_Snowflake_Uploader import upload_reset_SCH_WALMART_data
from Reset_Schedule_to_Snowflake_Uploader import upload_reset_SCH_RALEYS_data
from Reset_Schedule_to_Snowflake_Uploader import upload_reset_SCH_FOODMAXX_data
from Reset_Schedule_to_Snowflake_Uploader import upload_reset_SCH_SMART_FINAL_data
from Reset_Schedule_to_Snowflake_Uploader import upload_reset_SCH_SPROUTS_data
from Reset_Schedule_to_Snowflake_Uploader import upload_reset_SCH_TARGET_data
from Reset_Schedule_to_Snowflake_Uploader import upload_reset_SCH_WHOLEFOODS_data
from Reset_Schedule_to_Snowflake_Uploader import upload_reset_SCH_SAVEMART_data
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import datetime
from io import BytesIO
import numpy as np


    


#====================================================================================================================
# Setup page config and logo on the page
#====================================================================================================================

st.set_page_config(layout="wide", initial_sidebar_state="expanded")


def add_logo(logo_path, width, height):
    """Read and return a resized logo"""
    logo = Image.open(logo_path)
    modified_logo = logo.resize((width, height))
    return modified_logo

my_logo = add_logo(logo_path="./images/DeltaPacific_Logo.jpg", width=200, height = 100)
st.sidebar.image(my_logo)
st.sidebar.subheader("Delta Pacific Beverage Co.")




# Set Page Header   
st.header("CHAIN RESET MANAGMENT")
# Set custom CSS for hr element
st.markdown("""
        <style>
            hr {
                margin-top: 0.5rem;
                margin-bottom: 0.5rem;
                height: 3px;
                background-color: #333;
                border: none;
            }
        </style>
    """, unsafe_allow_html=True)

# Add horizontal line
st.markdown("<hr>", unsafe_allow_html=True)


#===================================================================================================================

# Create a container to hold the file uploader
#===================================================================================================================
file_container = st.container()





#====================================================================================================================
# CREATE CONNECTION TO SNOWFLAKE
#====================================================================================================================


# Load Snowflake credentials from the secrets.toml file
snowflake_creds = st.secrets["snowflake"]

# Establish a new connection to Snowflake
conn = snowflake.connector.connect(
    account=snowflake_creds["account"],
    user=snowflake_creds["user"],
    password=snowflake_creds["password"],
    warehouse=snowflake_creds["warehouse"],
    database=snowflake_creds["database"],
    schema=snowflake_creds["schema"]
)

#====================================================================================================================
# End CREATE CONNECTION TO SNOWFLAKE
#====================================================================================================================



#====================================================================================================================
# Block of code that handles the dropdown used to decide which chain the reset schedule if for
#====================================================================================================================

# Function to retrieve options from Snowflake table
def get_options():
    cursor = conn.cursor()
    cursor.execute('SELECT option_name FROM options_table ORDER BY option_name')
    options = [row[0] for row in cursor]
    return options

# Function to update options in Snowflake table
def update_options(options):
    cursor = conn.cursor()
    cursor.execute('DELETE FROM options_table')
    for option in options:
        cursor.execute("INSERT INTO options_table (option_name) VALUES (%s)", (option,))
    conn.commit()


# Retrieve options from Snowflake table
options = get_options()

# Initialize session state variables
if 'new_option' not in st.session_state:
    st.session_state.new_option = ""
if 'option_added' not in st.session_state:
    st.session_state.option_added = False

#===================================================================================================================
# # Option selection for sheet to format for which chain
#===================================================================================================================
with file_container:
    st.subheader(":blue[Reset Schedule File Format Uploader]")



# Check if options are available
    if not options:
        st.warning("No options available. Please add options to the list.")
    else:
        # Create the dropdown in Streamlit
        selected_option = st.selectbox(':red[Select the Chain Reset Schedule to format]', options + ['Add new option...'], key="existing_option")

    # Check if the selected option is missing and allow the user to add it
    if selected_option == 'Add new option...':
        st.write("You selected: Add new option...")
        
        # Show the form to add a new option
        with st.form(key='add_option_form', clear_on_submit=True):
            new_option = st.text_input('Enter the new option', value=st.session_state.new_option)
            submit_button = st.form_submit_button('Add Option')
            
            if submit_button and new_option:
                options.append(new_option)
                update_options(options)
                st.success('Option added successfully!')
                st.session_state.option_added = True

        # Clear the text input field
        st.session_state.new_option = ""
        
    else:
        # Display the selected option
        st.write(f"You selected: {selected_option}")

#===================================================================================================================
# END Option selection for sheet to format for which chain
#===================================================================================================================


    # File uploader for chain reset schedule spreadsheets
    uploaded_file = st.file_uploader(":red[Upload reset schedule spreadsheet to be formatted]", type=["xlsx"])


    formatted_workbook = None  # Initialize the variable

        

    if st.button("Reformat Spreadsheet"):
        with st.spinner('Starting Format of Spreadsheet ...'):
            if uploaded_file is None:
                st.warning("Please upload a spreadsheet first.")
            else:
                # Load the workbook
                workbook = openpyxl.load_workbook(uploaded_file)

            # Call the format_raleys_Schedule function for 'Raleys' option
            if selected_option == 'RALEYS':
                formatted_workbook = format_RALEYS_Schedule(workbook)
            # Call the format_SAFEWAY_Schedule function for 'Safeway' option
            elif selected_option == 'SAFEWAY':  # Add this condition for 'SAFEWAY' option
                formatted_workbook = format_SAFEWAY_Schedule(workbook)
            # Call the format_WALMART_Schedule function for 'Walmart' option
            elif selected_option == 'WALMART': #ADD THIS CONDITION FOR 'WALMART' OPTION
                formatted_workbook = format_WALMART_schedule(workbook)
            # Call the format_FOODMAXX_Schedule function for 'FoodMaxx' option
            elif selected_option == 'FOOD MAXX': #ADD THIS CONDITION FOR 'WALMART' OPTION
                formatted_workbook = format_FOODMAXX_schedule(workbook)
            # Call the format_LUCKYS_Schedule function for 'Luckys' option
            elif selected_option == 'LUCKYS': #ADD THIS CONDITION FOR 'LUCKYS' OPTION
                formatted_workbook = format_LUCKYS_Schedule(workbook)
            # Call the format_Savemart_Schedule function for 'Savemart' option
            elif selected_option == 'SAVEMART': #ADD THIS CONDITION FOR 'Save Mart' OPTION
                formatted_workbook = format_Savemart_Schedule(workbook)
            # Call the format_SPROUTS_Schedule function for 'Sprouts' option
            elif selected_option == 'SPROUTS': #ADD THIS CONDITION FOR 'SPROUTS' OPTION
                formatted_workbook = format_SPROUTS_Schedule(workbook)
            # Call the format_SMART_FINAL_Schedule function for 'Smart & Final' option
            elif selected_option == 'SMART_FINAL': #ADD THIS CONDITION FOR 'SMART & FINAL' OPTION
                 formatted_workbook =  format_SMARTFINAL_Schedule(workbook)

            else:
                # Call other formatting functions for different options
                formatted_workbook = workbook  # Use the original workbook

            # Create a new filename based on the selected option
            new_filename = f"formatted_{selected_option}_RESET_spreadsheet.xlsx"

            #Remove the autofilter if it exists
            if formatted_workbook is not None:
                for sheet_name in formatted_workbook.sheetnames:
                    sheet = formatted_workbook[sheet_name]
                    if sheet.auto_filter:
                        sheet.auto_filter.ref = None
                        st.write(f"Autofilter removed from '{sheet_name}' worksheet.")

    # Check if the workbook was successfully formatted
    if formatted_workbook is not None:
    
        # Save the formatted workbook to a stream
        stream = BytesIO()
        formatted_workbook.save(stream)
        stream.seek(0)

   
   
        # Provide the download link for the formatted spreadsheet
        st.download_button(
            label="Download formatted spreadsheet",
            data=stream.read(),
            file_name=new_filename,
            mime='application/vnd.ms-excel'
        )
        

    # Close the Snowflake connection
    conn.close()

#====================================================================================================================
# End Block of code that handles the dropdown used to decide which chain the reset schedule if for
#====================================================================================================================
        



# Add horizontal line
st.markdown("<hr>", unsafe_allow_html=True)


def update_metadata_table(table_name):
    cursor = conn.cursor()

    # Check if the record already exists in the metadata table
    cursor.execute("SELECT COUNT(*) FROM METADATA WHERE TABLE_NAME = %s", (table_name,))
    count = cursor.fetchone()[0]

    if count == 0:
        # Insert a new record into the metadata table
        cursor.execute("""
            INSERT INTO METADATA (TABLE_NAME, STATUS, CREATION_DATE)
            VALUES (%s, 'ACTIVE', CURRENT_TIMESTAMP())
        """, (table_name,))
    else:
        # Check if the table is being overwritten
        overwrite_confirmed = st.session_state.get('confirm_overwrite', False)

        if overwrite_confirmed:
            # Update the existing record in the metadata table
            cursor.execute("""
                UPDATE METADATA
                SET STATUS = 'ACTIVE', CREATION_DATE = CURRENT_TIMESTAMP()
                WHERE TABLE_NAME = %s
            """, (table_name,))

    # Commit the changes
    cursor.execute("COMMIT")

    # Close the cursor
    cursor.close()






def table_exists(conn, schema, table_name):
    cursor = conn.cursor()
    sql_query = f"SHOW TABLES IN {schema}"
    cursor.execute(sql_query)
    result = cursor.fetchall()
    cursor.close()
    existing_tables = [row[1] for row in result]
    return table_name.upper() in existing_tables


def create_table(conn, schema, table_name, column_names, column_types):
    cursor = conn.cursor()
    columns = ', '.join([f'"{col}" {col_type}' for col, col_type in zip(column_names, column_types)])
    sql_query = f'CREATE TABLE {schema}.{table_name} ({columns})'
    cursor.execute(sql_query)
    cursor.close()








# create file uploader
#uploaded_file = st.file_uploader(":red[Browse or selected formatted reset schedule excel sheet]", type=["xlsx"])

#===============================================================================================================================
# Utility to process formatted Reset Schedule to Snowflake depending on which chain you are working on
#===============================================================================================================================


with file_container:

    
        # Add horizontal line
    st.markdown("<hr>", unsafe_allow_html=True)
    st.subheader(":blue[Reset Schedule File to Upload to Snowflake Utility]")

    # Store the selected_option in session state
    if "selected_option" not in st.session_state:
        st.session_state.selected_option = None
    
    # Check if options are available
    if not options:
        st.warning("No options available. Please add options to the list.")
    else:
        # Create the dropdown in Streamlit
        selected_option = st.selectbox(':red[Select the Chain Reset Schedule to load to Snowflake]', options + ['Add new option...'], key="select_snowflake_option")

    # Check if the selected option is missing and allow the user to add it
    if selected_option == 'Add new option...':
        st.write("You selected: Add new option...")
        
        # Show the form to add a new option
        with st.form(key='add_option_form', clear_on_submit=True):
            new_option = st.text_input('Enter the new option', value=st.session_state.new_option)
            submit_button = st.form_submit_button('Add Option')
            
            if submit_button and new_option:
                options.append(new_option)
                update_options(options)
                st.success('Option added successfully!')
                st.session_state.option_added = True

        # Clear the text input field
        st.session_state.new_option = ""
        
    else:
        # Display the selected option
        st.write(f"You selected: {selected_option}")

    # Store selected_option in session state
    st.session_state.selected_option = selected_option



    # create file uploader
    uploaded_files = st.file_uploader(":red[Browse or select formatted reset schedule excel sheet To Upload to Snowflake]", type=["xlsx"], accept_multiple_files=True)

    # Process each uploaded file
    for uploaded_file in uploaded_files:
        # Read Excel file into pandas ExcelFile object
        excel_file = pd.ExcelFile(uploaded_file)

        # Get sheet names from ExcelFile object
        sheet_names = excel_file.sheet_names

        # Display workbook name and sheet names in Streamlit
        workbook_name = uploaded_file.name
        #st.write(f"Workbook Name: {workbook_name}")
        #st.write("Sheet Names:", sheet_names)

        ## Display DataFrame for each sheet in Streamlit
        for sheet_name in sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)

        #    # Modify DataFrame values directly to replace 'NAN' with empty string ''
            df = df.replace('NAN', np.nan)
            #st.write(df)

        #    

    # Write DataFrame to Snowflake on button click
        button_key = f"import_button_{workbook_name}_{sheet_name}"
        if st.button("Import into Snowflake", key=button_key):
            with st.spinner('Uploading data to Snowflake ...'):
                # Write DataFrame to Snowflake based on the selected store
                if selected_option == "SAFEWAY":
                    upload_reset_SCH_SAFEWAY_data(df, "COMPUTE_WH", "datasets", "DATASETS")
                elif selected_option == "LUCKYS":
                    upload_reset_SCH_LUCKY_data(df, "COMPUTE_WH", "datasets", "DATASETS")
                elif selected_option == "WALMART":
                    upload_reset_SCH_WALMART_data(df, "COMPUTE_WH", "datasets", "DATASETS")
                    # Add more if-else statements for other stores as needed
                elif selected_option == "RALEYS":
                    upload_reset_SCH_RALEYS_data(df, "COMPUTE_WH", "datasets", "DATASETS")
                    # Add more if-else statements for other stores as needed
                elif selected_option == "FOOD MAXX":
                    upload_reset_SCH_FOODMAXX_data(df, "COMPUTE_WH", "datasets", "DATASETS")
                elif selected_option == "SMART_FINAL":
                    upload_reset_SCH_SMART_FINAL_data(df, "COMPUTE_WH", "datasets", "DATASETS")
                elif selected_option == "SPROUTS":
                    upload_reset_SCH_SPROUTS_data(df, "COMPUTE_WH", "datasets", "DATASETS")
                elif selected_option == "TARGET":
                    upload_reset_SCH_TARGET_data(df, "COMPUTE_WH", "datasets", "DATASETS")
                elif selected_option == "WHOLEFOODS":
                    upload_reset_SCH_WHOLEFOODS_data(df, "COMPUTE_WH", "datasets", "DATASETS")
                elif selected_option == "SAVEMART":
                    upload_reset_SCH_SAVEMART_data(df, "COMPUTE_WH", "datasets", "DATASETS")
                    # Add more if-else statements for other stores as needed
                else:
                    # Call other formatting functions for different options
                    #st.write("test")#formatted_workbook = workbook  # Use the original workbook    
                    # Call other formatting functions for different options
                    formatted_workbook = workbook  # Use the original workbook
                    
                    
#======================================================================================================================
# This chunk of code takes the Supplier by county pivot table and formats in so it can be uploaded to snowflake
#=======================================================================================================================
#
def format_supplier_by_county(file_content):
    df_formatted = None

    # Perform data transformation
    df = pd.read_excel(file_content, sheet_name="Report")
    
    # Remove column "TOTAL" (column B)
    df = df.drop(columns=["TOTAL"])
    
    #Rename the column
    df.rename(columns={"Supplier / County": "Supplier"}, inplace=True)


    df_formatted = pd.melt(df, id_vars=["Supplier"], var_name="County", value_name="Status")

    # Change "Once" values to "Yes" or "No"
    df_formatted["Status"] = df_formatted["Status"].apply(lambda x: "Yes" if x == 1 else "No" if pd.isna(x) else x)

    return df_formatted



#===================================================================================================
# This function will write the Supplier County to a snowflake table for gap report county validation
#====================================================================================================
def write_to_snowflake(df_content):
    # Load Snowflake credentials from the secrets.toml file
    snowflake_creds = st.secrets["snowflake"]

    # Establish a new connection to Snowflake
    conn = snowflake.connector.connect(
        account=snowflake_creds["account"],
        user=snowflake_creds["user"],
        password=snowflake_creds["password"],
        warehouse=snowflake_creds["warehouse"],
        database=snowflake_creds["database"],
        schema=snowflake_creds["schema"]
    )
   
    
    # Create a cursor object to execute SQL queries
    cursor = conn.cursor()

    # Truncate the existing table before inserting new data (optional)
    truncate_query = "TRUNCATE TABLE DATASETS.DATASETS.SUPPLIER_COUNTY"
    cursor.execute(truncate_query)

    # Create an INSERT INTO query to insert the DataFrame data into the table
    insert_query = """
    INSERT INTO DATASETS.DATASETS.SUPPLIER_COUNTY (
        SUPPLIER,
        COUNTY,
        STATUS
    )
    VALUES (%s, %s, %s)
    """

    # Split the DataFrame into smaller batches
    batch_size = 100  # Adjust the batch size as per your needs
    batches = [df_content[i:i + batch_size] for i in range(0, len(df_content), batch_size)]

    # Iterate over each batch and execute batch insert
    for batch in batches:
        values = batch[["Supplier", "County", "Status"]].values.tolist()
        cursor.executemany(insert_query, values)

    # Commit the changes to the Snowflake table
    conn.commit()

    # Close the connection
    conn.close()

    
#===================================================================================================
# End of This function will write the Supplier County to a snowflake table for gap report county validation
#====================================================================================================
with file_container:
    st.markdown("<hr>", unsafe_allow_html=True)
    st.subheader(":blue[Process the Supplier by County Data and write to Snowflake]")
    # Create file uploader
    uploaded_file = st.file_uploader(":red[Upload Supplier by County Excel file]", type=["xlsx", "xls"])

    # Initialize session state
    if "df_formatted" not in st.session_state:
        st.session_state.df_formatted = None

    # Display the formatted DataFrame
    if uploaded_file is not None:
        # Show the Reformat button
        if st.button("Reformat Supplier by County Spreadsheet"):
            # Format the Supplier BY COUNTY excel spreadsheet
            file_content = uploaded_file.getvalue()  # Get the content of the uploaded file
            df_formatted = format_supplier_by_county(file_content)
            st.session_state.df_formatted = df_formatted
            new_csv_file = 'formatted_' + uploaded_file.name
            stream = BytesIO()
            df_formatted.to_excel(stream, index=False)
            stream.seek(0)
            st.download_button(label="Download formatted Supplier by County", data=stream.read(), file_name=new_csv_file, mime='application/vnd.ms-excel')
    
        


#=================================================================================================
# Process the formatted spreadsheet with Supplier by County data to snowflake
#=================================================================================================
    # uploaded_file = st.file_uploader(":red[Upload formated file to write to snowflake]", type=["xlsx", "xls"])
    

    # # Show the Write to Snowflake button
    # if st.button("Write Supplier by County Data to Snowflake"):
    #     if st.session_state.df_formatted is None:
    #         st.warning("Please click the 'Reformat Supplier by County Spreadsheet' button to format the data.")
    #     else:
    #         # Write the DataFrame data to Snowflake table
    #         st.write("Writing to Snowflake - df_formatted:")
    #         #st.write(st.session_state.df_formatted)
    #         write_to_snowflake(st.session_state.df_formatted)
            


uploaded_file_snowflake = st.file_uploader(":red[Upload formated file to write to snowflake]", type=["xlsx", "xls"])

# Show the Write to Snowflake button
if st.button("Write Supplier by County Data to Snowflake"):
    if st.session_state.df_formatted is None:
        st.warning("Please click the 'Reformat Supplier by County Spreadsheet' button to format the data.")
    elif uploaded_file_snowflake is None:
        st.warning("Please upload the formatted file to proceed.")
    else:
        # Load the formatted DataFrame from the session state
        df_formatted = st.session_state.df_formatted
        
        # Write the DataFrame data to Snowflake table
        st.write("Writing to Snowflake - df_formatted:")
        #st.write(df_formatted)  # Display the formatted data (optional)
        
        # Write the data to Snowflake
        write_to_snowflake(df_formatted)
        st.success("Data written to Snowflake successfully!")



