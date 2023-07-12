# Import required libraries
from importlib.resources import Package
from multiprocessing.connection import wait
from this import s
import snowflake.connector
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import numpy as np
from io import BytesIO
from openpyxl import Workbook
import datetime
from PIL import Image
import streamlit.components.v1 as components



#=====================================================================================================================================
#Page configuration and page formatting
#=====================================================================================================================================

st.set_page_config(layout="wide")
padding_top = 10

#reduces the space above the header so the header will be at the top of the page
reduce_header_height_style = """
    <style>
        div.block-container {padding-top:1rem;}
    </style>
"""
st.markdown(reduce_header_height_style, unsafe_allow_html=True)





def add_logo(logo_path, width, height):
    """Read and return a resized logo"""
    logo = Image.open(logo_path)
    modified_logo = logo.resize((width, height))
    return modified_logo

my_logo = add_logo(logo_path="./images/DeltaPacific_Logo.jpg", width=200, height = 100)
st.sidebar.image(my_logo)
st.sidebar.subheader("Delta Pacific Beverage Co.")



# Set Page Header
st.header("Load Customer and Products Table Data")

# Set custom CSS for hr element
st.markdown("""
        <style>
            hr {
                margin-top: 0.5rem;
                margin-bottom: 0.5rem;
                height: 3px;
                background-color: #333;
                border: none;
            }
        </style>
    """, unsafe_allow_html=True)

# Add horizontal line
st.markdown("<hr>", unsafe_allow_html=True)

#=====================================================================================================================================
# End Page configuration and page formatting
#=====================================================================================================================================


#=========================================================================================================================

# The following code with reformat the Customers Excel worksheet in to importable Data
#=========================================================================================================================
def format_Customers_Dataload(workbook):
    # Delete all sheets except Customers
    for sheet_name in workbook.sheetnames:
        if sheet_name != 'Customers':
            workbook.remove(workbook[sheet_name])

    # Select the Customers sheet
    ws = workbook['Customers']

    # Remove filters
    ws.auto_filter.ref = None

    # Create a new column for store name
    ws.insert_cols(3)
    #ws.cell(row=1, column=3, value='STORE NAME')
    
    # Copy values before the # to store name column
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if '#' in str(cell.value):
                cell_offset = ws.cell(row=cell.row, column=4)
                store_name = str(cell.value).split('#')[0].replace("#", "")
                cell_offset.value = store_name



    # Move values from column D to column C
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell_offset = ws.cell(row=cell.row, column=3)
            cell_offset.value = cell.value

    # Delete original column D (now column E)
    ws.delete_cols(5)
   
    #ws.cell(row=1, column=3, value='Store_Number')


    # Rename Columns as required to meet objective for uploading to Snowflake
    ws.cell(row=1, column=1, value='Customer_id')
    ws.cell(row=1, column=2, value='Chain_Name')
    ws.cell(row=1, column=3, value='Store_Number')
    ws.cell(row=1, column=4, value='Store_Name')
    ws.cell(row=1, column=5, value='Address')
    ws.cell(row=1, column=6, value='City')
    ws.cell(row=1, column=7, value='County')
    ws.cell(row=1, column=8, value='Salesperson')
    ws.cell(row=1, column=9, value='Account_Status')
    


    # Remove all apostrophes (') from column E
    for cell in ws['E']:
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.replace("'", "")

    # Remove all apostrophes (') from column B 
    for cell in ws['B']:
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.replace("'", "")



   



    return workbook


#-------------------------------------------------------------------------------------------------------------------------------------
# Handles page look, creates uploader and once uploaded provides button to call the format function

# Make the text red
st.markdown('<p style="color: red;">Upload Customer Report from Encompass</p>', unsafe_allow_html=True)

# Upload the workbook
uploaded_file = st.file_uploader("Browse or Select File", type=["xlsx", "xls"])

if uploaded_file is not None:
    # Load the workbook
    workbook = openpyxl.load_workbook(uploaded_file)

    # Show the Reformat button upon selecting will call the format_Customers_dataload function
    if st.button("Reformat"):
        # Format the sales report
        new_workbook = format_Customers_Dataload(workbook)

        # Download the formatted file
        new_filename = 'formatted_' + uploaded_file.name
        stream = BytesIO()
        new_workbook.save(stream)
        stream.seek(0)
        st.download_button(label="Download formatted file", data=stream.read(), file_name=new_filename, mime='application/vnd.ms-excel')
    


#=====================================================================================================================================

# The below function formats the Products report to prepare for upload to snowflake table
#=====================================================================================================================================


def format_Products_Dataload(workbook):
    # Delete all sheets except Products
    for sheet_name in workbook.sheetnames:
        if sheet_name != 'Products':
            workbook.remove(workbook[sheet_name])

    # Select the Products sheet
    ws = workbook['Products']

    
    # Create a new column for Manufacturer
    ws.insert_cols(2)
    ws.cell(row=1, column=2, value='Manufacturer')

  


    # Create a new column for Package
    ws.insert_cols(4)
    ws.cell(row=1, column=4, value='Package')


       # Create a new column for UPC
    ws.insert_cols(5)
    ws.cell(row=1, column=5, value='UPC')

       # Move values from column H to column B
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
        for cell in row:
            cell_offset = ws.cell(row=cell.row, column=2)
            cell_offset.value = cell.value



    # Rename column A to Customer_id
    ws['A1'] = 'Product_id'

    # Move values from column F to column B
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            cell_offset = ws.cell(row=cell.row, column=4)
            cell_offset.value = cell.value

 
    # Move values from column D to column C
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell_offset = ws.cell(row=cell.row, column=5)
            cell_offset.value = cell.value
            cell_offset.value = cell.value

    #Delete column G
    ws.delete_cols(7)

    #Delete column H
    ws.delete_cols(8)

    #Delete column H
    ws.delete_cols(9)

    #Delete column H
    ws.delete_cols(7)


   
    # Remove all Hyphens in column B
    for cell in ws['B']:
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.replace(',', ' ')

    # Remove all Hyphens in column E
    for cell in ws['C']:
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.replace('\'', "")

    # Replace all commas with spaces in column C
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.replace(',', ' ')


   
          
    # Remove all Is Null from column E
    for cell in ws['E']:
        if cell.value is None:
            cell.value = 99


    

    return workbook


# Add horizontal line
st.markdown("<hr>", unsafe_allow_html=True)
# Make the text red
st.markdown('<p style="color: red;">Upload Products Report from Encompass</p>', unsafe_allow_html=True)


# Upload the workbook
uploaded_file = st.file_uploader("Browse or select File", type=["xlsx", "xls"])

if uploaded_file is not None:
    # Load the workbook
    workbook = openpyxl.load_workbook(uploaded_file)

    # Show the Reformat button
    if st.button("Reformat_Products"):
        # Format the sales report
        new_workbook = format_Products_Dataload(workbook)

        # Download the formatted file
        new_filename = 'formatted_' + uploaded_file.name
        stream = BytesIO()
        new_workbook.save(stream)
        stream.seek(0)
        st.download_button(label="Download formatted file", data=stream.read(), file_name=new_filename, mime='application/vnd.ms-excel')
    




def write_to_products_snowflake(df, warehouse, database, schema):
    
    
    # read Excel file into pandas DataFrame
    df = pd.read_excel(uploaded_file)
    
   
    
    # Modify DataFrame values directly to replace 'NAN' with empty string ''
    
    df = df.replace('NAN', np.nan)
    
    # Load Snowflake credentials from the secrets.toml file
    snowflake_creds = st.secrets["snowflake"]

    # Establish a new connection to Snowflake
    conn = snowflake.connector.connect(
    account=snowflake_creds["account"],
    user=snowflake_creds["user"],
    password=snowflake_creds["password"],
    warehouse=snowflake_creds["warehouse"],
    database=snowflake_creds["database"],
    schema=snowflake_creds["schema"]
    )
    
    
    
    
   # write DataFrame to Snowflake
    cursor = conn.cursor()
    
    # Generate the SQL query
    values = ', '.join([str(tuple(row)) for row in df.values])
    sql_query = f"""
    CREATE OR REPLACE TABLE PRODUCTS AS
    SELECT
        CAST(PRODUCT_ID AS NUMBER) AS PRODUCT_ID,
        CAST(SUPPLIER AS VARCHAR) AS SUPPLIER,
        CAST(PRODUCT_NAME AS VARCHAR) AS PRODUCT_NAME,
        CAST(PACKAGE AS VARCHAR) AS PACKAGE,
        CARRIER_UPC,
        CAST(PRODUCT_MANAGER AS VARCHAR) AS PRODUCT_MANAGER
    FROM
        (VALUES {values}) AS tmp(
            PRODUCT_ID,
            SUPPLIER,
            PRODUCT_NAME,
            PACKAGE,
            UPC,
            PRODUCT_MANAGER
        );
    """






# Add horizontal line
st.markdown("<hr>", unsafe_allow_html=True)
# Make the text red
st.markdown('<p style="color: red;">UPLOAD PRODUCTS REPORT AFTER IT HAS BEEN FORMATED</p>', unsafe_allow_html=True)

# create file uploader
uploaded_file = st.file_uploader("Browse or Select File", type=["xlsx"])


# check if file was uploaded
if uploaded_file:
    # read Excel file into pandas DataFrame
    df = pd.read_excel(uploaded_file)
    #print(df.columns)
    # display DataFrame in Streamlit
    st.dataframe(df)


    # write DataFrame to Snowflake on button click
    if st.button("Import into Snowflake"):
        with st.spinner('Uploading product data to Snowflake ...'):
            write_to_products_snowflake(df, "COMPUTE_WH", "datasets", "DATASETS")
            st.write("Product Data has been imported into Snowflake table! ", "Products")






#====================================================================================================================================================

# Writes Customer table data to the customers  table in snowflake
#----------------------------------------------------------------------------------------------------------------------------------------------------
def write_to_customers_snowflake(df, warehouse, database, schema):
    # Modify DataFrame values directly to replace 'NAN' with empty string ''
    df = df.replace('NAN', np.nan)

    # Load Snowflake credentials from the secrets.toml file
    snowflake_creds = st.secrets["snowflake"]

    # Establish a new connection to Snowflake
    conn = snowflake.connector.connect(
        account=snowflake_creds["account"],
        user=snowflake_creds["user"],
        password=snowflake_creds["password"],
        warehouse=snowflake_creds["warehouse"],
        database=snowflake_creds["database"],
        schema=snowflake_creds["schema"]
    )

    # write DataFrame to Snowflake
    cursor = conn.cursor()

    # Generate the SQL query
    values = ', '.join([str(tuple(row)) for row in df.values])
    sql_query = f"""
    CREATE OR REPLACE TABLE CUSTOMERS AS
    SELECT
        CAST(CUSTOMER_ID AS NUMBER) AS CUSTOMER_ID,
        CAST(CHAIN_NAME AS VARCHAR) AS CHAIN_NAME,
        CAST(STORE_NUMBER AS NUMBER) AS STORE_NUMBER,
        CAST(STORE_NAME AS VARCHAR) AS STORE_NAME,
        CAST(ADDRESS AS VARCHAR) AS ADDRESS,
        CAST(CITY AS VARCHAR) AS CITY,
        CAST(COUNTY AS VARCHAR) AS COUNTY,
        CAST(SALESPERSON AS VARCHAR) AS SALESPERSON,
        CAST(ACCOUNT_STATUS AS VARCHAR) AS ACCOUNT_STATUS
    FROM
        (VALUES {values}) AS tmp(
            CUSTOMER_ID,
            CHAIN_NAME,
            STORE_NUMBER,
            STORE_NAME,
            ADDRESS,
            CITY,
            COUNTY,
            SALESPERSON,
            ACCOUNT_STATUS
        );
    """

    # Execute the SQL query
    cursor.execute(sql_query)

    # Commit the changes to the Snowflake table
    conn.commit()

    # Close the cursor and the connection
    cursor.close()
    conn.close()


# Add horizontal line
st.markdown("<hr>", unsafe_allow_html=True)
# Make the text red
st.markdown('<p style="color: red;">UPLOAD CUSTOMERS REPORT AFTER IT HAS BEEN FORMATTED</p>', unsafe_allow_html=True)

# create file uploader
uploaded_file = st.file_uploader("Browse or Select File1", type=["xlsx"])

# check if file was uploaded
if uploaded_file:
    # read Excel file into pandas DataFrame
    df = pd.read_excel(uploaded_file)
    # display DataFrame in Streamlit
    st.dataframe(df)

    # write DataFrame Customers to Snowflake Customers Table on button click
    if st.button("Import into Snowflake"):
        with st.spinner('Uploading Customers data to Snowflake ...'):
            write_to_customers_snowflake(df, "COMPUTE_WH", "datasets", "DATASETS")
            st.write("Customer Data has been imported into Snowflake table! ", "Customers")

