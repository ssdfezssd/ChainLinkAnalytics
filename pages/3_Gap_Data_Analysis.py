# Import required libraries
from cgitb import text
from pickle import TRUE
from this import s
import snowflake.connector
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import numpy as np
from io import BytesIO
import os
from openpyxl import Workbook
import datetime
import base64
import plotly.express as px
import altair as alt
import plotly.graph_objects as go
from snowflake.connector import connect, Error



# Displaying images on the front end
from PIL import Image
st.set_page_config(layout="wide")


def add_logo(logo_path, width, height):
    """Read and return a resized logo"""
    logo = Image.open(logo_path)
    modified_logo = logo.resize((width, height))
    return modified_logo

my_logo = add_logo(logo_path="./images/DeltaPacific_Logo.jpg", width=200, height = 100)
st.sidebar.image(my_logo)
st.sidebar.subheader("Delta Pacific Beverage Co.")
st.subheader("Gap Report and Analysis")

# Set custom CSS for hr element
st.markdown("""
        <style>
            hr {
                margin-top: 0.5rem;
                margin-bottom: 0.5rem;
                height: 3px;
                background-color: #333;
                border: none;
            }
        </style>
    """, unsafe_allow_html=True)

# Add horizontal line
st.markdown("<hr>", unsafe_allow_html=True)




# Allow the user to select the environment
ENVIRONMENT = st.selectbox(
    'Select environment:',
    ('PRODUCTION', 'TEST')
    )

# Set up a session state for storing table_name based on the environment
if ENVIRONMENT == 'PRODUCTION':
   st.session_state.table_name = 'SALES_REPORT'
else:
    st.session_state.table_name = 'TMP_TABLE'

#st.write('you have selected the ' + ENVIRONMENT) # Use this line for testing which environment you are selecting




def format_sales_report(workbook):
    # Delete all sheets except SALES REPORT
    for sheet_name in workbook.sheetnames:
        if sheet_name != 'SALES REPORT':
            workbook.remove(workbook[sheet_name])

    # Select the SALES REPORT sheet
    ws = workbook['SALES REPORT']

    # Delete row 2
    ws.delete_rows(2)

    # Delete column H
    ws.delete_cols(8)

    # Remove all hyphens from column F
    for cell in ws['F']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('-', '')

    # Create a new column for store name
    ws.insert_cols(2)
    ws.cell(row=1, column=2, value='STORE NAME')

    # Copy values before the # to store name column
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if '#' in str(cell.value):
                cell_offset = ws.cell(row=cell.row, column=2)
                store_name = str(cell.value).split('#')[0].replace("'", "")
                cell_offset.value = store_name

    # Remove column C
    ws.delete_cols(3)

    # Replace all commas with spaces in column B
    for cell in ws['B']:
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.replace(',', ' ')

    # Remove all 's in column B
    for cell in ws['B']:
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.replace(" 's", "")

    # Replace all commas with spaces in column E
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.replace(',', ' ')

       # Replace all single quote with spaces in column E
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.replace("'", ' ')

   
    # Remove all commas from column C
    for cell in ws['C']:
        if cell.value is not None:
            cell.value = str(cell.value).replace(',', ' ')
            
            
            
    # Remove all Is Null from column F
    for cell in ws['F']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('Is Null', '0')

    # Format column G as number with no decimals
    for cell in ws['G'][1:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = numbers.FORMAT_NUMBER
        elif isinstance(cell.value, str):
            cell.number_format = numbers.FORMAT_NUMBER
            try:
                cell.value = float(cell.value.replace(",", ""))
            except ValueError:
                pass
                
    
    

    return workbook



# Upload the workbook
uploaded_file = st.file_uploader(":red[Upload freshly ran sales report from Encompass]", type=["xlsx", "xls"])

if uploaded_file is not None:
    # Load the workbook
    workbook = openpyxl.load_workbook(uploaded_file)

    # Show the Reformat button
    if st.button("Reformat"):
        # Format the sales report
        new_workbook = format_sales_report(workbook)

        # Download the formatted file
        new_filename = 'formatted_' + uploaded_file.name
        stream = BytesIO()
        new_workbook.save(stream)
        stream.seek(0)
        st.download_button(label="Download formatted file", data=stream.read(), file_name=new_filename, mime='application/vnd.ms-excel')
    

#=====================================================================================================
# Function to write sales report data to snowflake
#=====================================================================================================

def write_salesreport_to_snowflake(df, warehouse, database, schema, table_name):
    
    
    # read Excel file into pandas DataFrame
    df = pd.read_excel(uploaded_file)
    
   
    # replace NaN values with "NULL"
    df.fillna(value=np.nan, inplace=True)
    



 
    # Load Snowflake credentials from the secrets.toml file
    snowflake_creds = st.secrets["snowflake"]

    # Establish a new connection to Snowflake
    conn = snowflake.connector.connect(
    account=snowflake_creds["account"],
    user=snowflake_creds["user"],
    password=snowflake_creds["password"],
    warehouse=snowflake_creds["warehouse"],
    database=snowflake_creds["database"],
    schema=snowflake_creds["schema"]
    )
    
    
    
 
    # write DataFrame to Snowflake
    cursor = conn.cursor()
    sql_query = f"CREATE OR REPLACE TABLE {table_name} AS SELECT \
    CAST(STORE_NUMBER AS NUMBER) AS STORE_NUMBER, \
    CAST(STORE_NAME AS VARCHAR) AS STORE_NAME, \
    CAST(ADDRESS AS VARCHAR) AS ADDRESS, \
    CAST(SALESPERSON AS VARCHAR) AS SALESPERSON, \
    CAST(PRODUCT_NAME AS VARCHAR) AS PRODUCT_NAME, \
    CAST(UPC AS VARCHAR) AS UPC, \
    CAST(PURCHASED_YES_NO AS VARCHAR) AS PURCHASED_YES_NO \
    FROM (VALUES {', '.join([str(tuple(df.iloc[i].fillna(np.nan).values)) for i in range(len(df))])}) \
    AS tmp(STORE_NUMBER, STORE_NAME, ADDRESS, SALESPERSON, PRODUCT_NAME, UPC, PURCHASED_YES_NO);"

    #st.write(sql_query)  # print the SQL query
    cursor.execute(sql_query)
    cursor.close()
    conn.close()
    st.write("Data has been imported into Snowflake table! ",st.session_state.table_name)



#=====================================================================================================
# Function to write sales report data to snowflake
#=====================================================================================================

#=====================================================================================================
# Create uploader for formatted sales report create dataframe and call write to snowflake function
#=====================================================================================================
# create file uploader
uploaded_file = st.file_uploader(":red[UPLOAD CURRENT SALES REPORT AFTER IT HAS BEEN FORMATED]", type=["xlsx"])

# check if file was uploaded
if uploaded_file:
    # read Excel file into pandas DataFrame
    df = pd.read_excel(uploaded_file)
    print(df.columns)
    # display DataFrame in Streamlit
    st.dataframe(df)

    # get warehouse and schema name from user
   
    #print(df.columns)

    # write DataFrame to Snowflake on button click
    if st.button("Import into Snowflake"):
        with st.spinner('Uploading data to Snowflake ...'):
            write_salesreport_to_snowflake(df, "COMPUTE_WH", "datasets", "DATASETS", st.session_state.table_name)

#=====================================================================================================
# END Create uploader for formatted sales report create dataframe and call write to snowflake function
#=====================================================================================================

#===================================================================================================
# Function to create the gap report from data pulled from snowflake and button to download gap report
#=====================================================================================================




def create_gap_report(conn, salesperson, store, supplier):
    """
    Retrieves data from a Snowflake view and creates a button to download the data as a CSV report.
    """
    # Execute the stored procedure without filters
    cursor = conn.cursor()
    cursor.execute("CALL PROCESS_GAP_REPORT()")
    cursor.close()

    # Execute SQL query and retrieve data from the Gap_Report view with filters
    if salesperson != "All":
        query = f"SELECT * FROM Gap_Report WHERE SALESPERSON = '{salesperson}'"
        if store != "All":
            query += f" AND STORE_NAME = '{store}'"
            if supplier != "All":
                query += f" AND SUPPLIER = '{supplier}'"
    elif store != "All":
        query = f"SELECT * FROM Gap_Report WHERE STORE_NAME = '{store}'"
        if supplier != "All":
            query += f" AND SUPPLIER = '{supplier}'"
    else:
        if supplier != "All":
            query = f"SELECT * FROM Gap_Report WHERE SUPPLIER = '{supplier}'"
        else:
            query = "SELECT * FROM Gap_Report"
    df = pd.read_sql(query, conn)

    # Get the user's download folder
    download_folder = os.path.expanduser(r"~\Downloads")

    # Write the updated dataframe to a temporary file
    temp_file_name = 'temp.xlsx'

    # Create the full path to the temporary file
    #temp_file_path = os.path.join(download_folder, temp_file_name)
    temp_file_path = "temp.xlsx"
    #df.to_excel(temp_file_path, index=False)

    df.to_excel(temp_file_path, index=False)  # Save the DataFrame to a temporary file


    # # Create the full path to the temporary file
    # temp_file_name = 'temp.xlsx'
    # temp_file_path = os.path.join(download_folder, temp_file_name)

    return temp_file_path  # Return the file path

#====================================================================================================
# Build sidebar button for creating gap report and call function to create the gap report
#====================================================================================================
# Load Snowflake credentials from the secrets.toml file
snowflake_creds = st.secrets["snowflake"]

# Establish a new connection to Snowflake
conn = snowflake.connector.connect(
    account=snowflake_creds["account"],
    user=snowflake_creds["user"],
    password=snowflake_creds["password"],
    warehouse=snowflake_creds["warehouse"],
    database=snowflake_creds["database"],
    schema=snowflake_creds["schema"]
)

# Retrieve salesperson, store, and supplier data from tables
salesperson_options = ["All"] + pd.read_sql("SELECT DISTINCT SALESPERSON FROM Salesperson", conn)['SALESPERSON'].tolist()
store_options = ["All"] + pd.read_sql("SELECT DISTINCT STORE_NAME FROM CUSTOMERS", conn)['STORE_NAME'].tolist()
supplier_options = ["All"]+ pd.read_sql("SELECT DISTINCT SUPPLIER FROM SUPPLIER_COUNTY", conn)['SUPPLIER'].tolist()

#  Create a form in the sidebar
with st.sidebar.form(key="Gap Report Report", clear_on_submit=True):
    # Select boxes for filters
    salesperson = st.selectbox("Filter by Salesperson", salesperson_options)
    store = st.selectbox("Filter by Chain", store_options)
    supplier = st.selectbox("Filter by Supplier", supplier_options)

    # Submit button
    submitted = st.form_submit_button("Generate Gap Report")

# Define the df variable outside of the function
df = None




with st.sidebar:
   # Call create_gap_report if form is submitted
   if submitted:
        with st.spinner('Generating report...'):
            temp_file_path = create_gap_report(conn, salesperson=salesperson, store=store, supplier=supplier)
            with open(temp_file_path, 'rb') as f:
                bytes_data = f.read()
            today = datetime.datetime.today().strftime('%Y-%m-%d') # get current date in YYYY-MM-DD format
            file_name = f"Gap_Report_{today}.xlsx" # insert current date into file name

            downloadcontainer = st.container()
            with downloadcontainer:
                st.download_button(label="Download Gap Report", data=bytes_data, file_name=file_name, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                st.write("File will be downloaded to your local download folder")

            container = st.container()
            with container:
                st.spinner('Generating report...')  # Display the spinner in the sidebar
                #st.sidebar.dataframe(df)  # Display the dataframe in the sidebar

   







#====================================================================================================
# END Build sidebar button for creating gap report and call function to create the gap report
#====================================================================================================

#===================================================================================================
# Gap Analysis Bar Chart code
#===================================================================================================

# Establish a new connection to Snowflake
conn = snowflake.connector.connect(
    account=snowflake_creds["account"],
    user=snowflake_creds["user"],
    password=snowflake_creds["password"],
    warehouse=snowflake_creds["warehouse"],
    database=snowflake_creds["database"],
    schema=snowflake_creds["schema"]
)

# Retrieve data from your view
query = "SELECT SUM(\"In_Schematic\") AS total_in_schematic, SUM(\"PURCHASED_YES_NO\") AS purchased, SUM(\"PURCHASED_YES_NO\") / COUNT(*) AS purchased_percentage FROM GAP_REPORT;"
df = pd.read_sql(query, conn)

# Format the 'PURCHASED_PERCENTAGE' column as a percentage with 2 decimal places using apply() function
df['PURCHASED_PERCENTAGE'] = (df['PURCHASED'] / df['TOTAL_IN_SCHEMATIC'] * 100).map('{:.2f}%'.format)

# Create the bar chart
fig = go.Figure(data=[
    go.Bar(x=['Total in Schematic', 'Purchased', 'Purchased Percentage'], 
           y=[df['TOTAL_IN_SCHEMATIC'].iloc[0], df['PURCHASED'].iloc[0], df['PURCHASED_PERCENTAGE'].iloc[0]], 
           text=[df['TOTAL_IN_SCHEMATIC'].iloc[0], df['PURCHASED'].iloc[0], df['PURCHASED_PERCENTAGE'].iloc[0]], 
           textposition='auto', 
           marker=dict(color=['#1f77b4', '#ff7f0e', '#2ca02c']))
])

# Set the axis labels and plot title
fig.update_layout(
    xaxis_title='',
    yaxis_title='Number of Items',
    title='Total Items in Schematic vs. Purchased Items',
    plot_bgcolor='#f2f2f2',  # Set the background color
    paper_bgcolor='#d9d9d9'  # Set the paper (border) color
)

# Add a border to the chart
fig.update_traces(
    marker_line_width=1.5,
    marker_line_color='black'
)

# Customize the bar chart colors
colors = ['#1f77b4', '#ff7f0e', '#2ca02c']
for i in range(len(fig.data)):
    fig.data[i].marker.color = colors[i]
    fig.data[i].marker.line.width = 1.5
    fig.data[i].marker.line.color = 'black'

# Row A
col1, col2  = st.columns(2)

with col1:
    container = st.container()
    with container:
        st.plotly_chart(fig, use_container_width=True)

with col2:
    container = st.container()
    with container:
        st.plotly_chart(fig, use_container_width=True)

   

#==================================================================================================
# END Gap Analysis Bar Chart code
#===================================================================================================