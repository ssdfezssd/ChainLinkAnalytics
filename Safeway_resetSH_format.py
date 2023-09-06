import streamlit as st
from openpyxl.styles import Border, Side, PatternFill, Font, NamedStyle
import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


#===================================================================================================================
# Reformat Excel spreadsheet for SAFEWAY
#====================================================================================================================
def format_SAFEWAY_Schedule(workbook):

    st.write("YAY YOU CALLED ME SAFEWAY")
    
    # # Delete all sheets except Reset Dates
    # for sheet_name in workbook.sheetnames:
    #     if sheet_name != 'Mainland':
    #             workbook.remove(workbook[sheet_name])

    # Select the Reset Dates sheet
    ws = workbook['Mainland']
    #remove_hidden_rows_and_columns(ws)

    # Remove filter from the worksheet
    ws.auto_filter.ref = None


        # Trim leading and trailing spaces in all cells
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):  # Check if the cell value is a string
                cell.value = cell.value.strip()

    
    # Determine the number of rows with data
    sheet_name = "Mainland"  # Update with the actual sheet name
    ws = workbook[sheet_name]
    max_row = ws.max_row
    for row in reversed(range(1, max_row + 1)):
        if ws.cell(row=row, column=1).value:
            max_row = row
            break

   # Delete extra rows beyond the data range
    if ws.max_row > max_row:
        ws.delete_rows(max_row + 1, ws.max_row - max_row)

    #st.write("the number of rows are", " ", max_row)

    # Get the maximum row number in column A
    max_row_a = ws.max_row

    # Create new column C to hold store name
    ws.insert_cols(3)

   

     # Create new column d to hold phone number
    ws.insert_cols(4)

    # Create new column E to hold city
    ws.insert_cols(5)

     # Create new column F to hold address
    ws.insert_cols(6)

   

    # Create new column H to hold store number
    ws.insert_cols(7)



    

    # Add "SAFEWAY" to each cell in column A
    for row in range(1, max_row_a + 1):
        cell = ws.cell(row=row, column=1)
        cell.value = "SAFEWAY"

    # Get the maximum row number in column C
    max_row_c = ws.max_row

    # Add "SAFEWAY" to each cell in column C
    for row in range(1, max_row_c + 1):
        cell = ws.cell(row=row, column=3)
        cell.value = "SAFEWAY"

    #st.write("Now the number of rows are", " ", max_row)

    # Iterate over the rows in column I and copy the values to column E
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
        for cell in row:
            # Get the value from column I
            value = cell.value

            # Calculate the destination column index (E = 5)
            dest_col = 5

            # Get the destination cell in column E
            dest_cell = ws.cell(row=cell.row, column=dest_col)

            # Set the value in the destination cell
            dest_cell.value = value

            # Clear the original cell in column I
            cell.value = None

    # Iterate over the rows in column H and copy the values to column F
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            # Get the value from column H
            value = cell.value

            # Calculate the destination column index (F = 6)
            dest_col = 6

            # Get the destination cell in column E
            dest_cell = ws.cell(row=cell.row, column=dest_col)

            # Set the value in the destination cell
            dest_cell.value = value

            # Clear the original cell in column I
            cell.value = None

     # Iterate over the rows in column J and copy the values to column G
    for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):
        for cell in row:
            # Get the value from column H
            value = cell.value

            # Calculate the destination column index (G = 7)
            dest_col = 7

            # Get the destination cell in column E
            dest_cell = ws.cell(row=cell.row, column=dest_col)

            # Set the value in the destination cell
            dest_cell.value = value

            # Clear the original cell in column I
            cell.value = None

    # Iterate over the rows in column L and copy the values to column J
    for row in ws.iter_rows(min_row=2, min_col=12, max_col=12):
        for cell in row:
            # Get the value from column L
            value = cell.value

            # Calculate the destination column index (J = 10)
            dest_col = 10

            # Get the destination cell in column J
            dest_cell = ws.cell(row=cell.row, column=dest_col)

            # Set the value in the destination cell
            dest_cell.value = value

            # Clear the original cell in column I
            cell.value = None

    
   # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, min_col=11), start=2):
        # Clear the cell value in column 
        ws.cell(row=index, column=11).value = None


# Iterate over the rows in column M and copy the values to column K
    for row in ws.iter_rows(min_row=2, min_col=13, max_col=13):
        for cell in row:
            # Get the value from column K
            value = cell.value

            # Calculate the destination column index (K = 11)
            dest_col = 11

            # Get the destination cell in column K
            dest_cell = ws.cell(row=cell.row, column=dest_col)

            # Set the value in the destination cell
            dest_cell.value = value

            # Clear the original cell in column M
            cell.value = None


  
# Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        # Specify the column you want to format (e.g., column A)
        column_letter = 'J'

    # Iterate over the cells in the column and set the number format
    for cell in ws[column_letter]:
        cell.number_format = 'mm/dd/yyyy'



 
    # Iterate over the rows and automatically adjust the row height
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                ws.row_dimensions[cell.row].height = 40  # specify the desired row height

    # Save the workbook
    #workbook.save("your_workbook.xlsx")

    

    # Specify the column to update (e.g., column K)
    column = 'K'

    # Iterate over the cells in the column and update the values
    for cell in ws[column]:
        if isinstance(cell.value, str):  # Check if the cell value is a string
            value = cell.value.strip().upper()  # Remove leading/trailing spaces and convert to uppercase
            if value == '6AM':
                cell.value = '6:00'
            elif value == '5AM':
                cell.value = '5:00'
            elif value == '7AM':
                cell.value = '7:00'
            elif value == '8AM':
                cell.value = '8:00'

    # Apply the desired time format
    for cell in ws[column]:
        if isinstance(cell.value, (datetime.datetime, datetime.time)):  # Check if the cell value is a datetime or time object
            cell.number_format = 'h:mm AM/PM'

  
    # Apply gridlines to all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

            # Remove cell colors
            cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

    # Make the entire sheet have black-colored text with font size 11
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(color="00000000", size=11)


    # Rename Columns as required to meet objective for uploading to Snowflake
    ws.cell(row=1, column=1, value='CHAIN_NAME')
    ws.cell(row=1, column=2, value='STORE_NUMBER')
    ws.cell(row=1, column=3, value='STORE_NAME')
    ws.cell(row=1, column=4, value='PHONE')
    ws.cell(row=1, column=5, value='CITY')
    ws.cell(row=1, column=6, value='ADDRESS')
    ws.cell(row=1, column=7, value='STATE')
    ws.cell(row=1, column=8, value='COUNTY')
    ws.cell(row=1, column=9, value='TEAM_LEAD')
    ws.cell(row=1, column=10, value='RESET_DATE')
    ws.cell(row=1, column=11, value='RESET_TIME')
    ws.cell(row=1, column=12, value='STATUS')
    ws.cell(row=1, column=13, value='NOTES')



    return workbook

    


#===================================================================================================================
# End Reformat Excel spreadsheet for SAFEWAY

#====================================================================================================================