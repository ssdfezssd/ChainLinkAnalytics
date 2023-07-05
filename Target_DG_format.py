import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import openpyxl.utils.datetime as xl_datetime
import numpy as np
from io import BytesIO
from openpyxl import Workbook


# Reformat Excel spreadsheet for SAFEWAY
#====================================================================================================================
def format_TARGET_DistroGrid(workbook):

    st.write("YAY YOU CALLED ME TARGET dg CODE")


    # Select the Reset Dates sheet
    ws = workbook['TARGET']

    # Insert a new column "Chain_Name" at the beginning and fill it with "WALMART"
    ws.insert_cols(1)
    ws.cell(row=1, column=1, value="STORE_NAME")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.value = "TARGET"

    # Delete column D 'ITEMLOC'
    ws.delete_cols(4)  

    # Delete column D 'DPCI' this becomes D after the delete above so not deleting the same thing here
    ws.delete_cols(4)  

    # insert new column F or column 6 for SKU
    ws.insert_cols(5)

    # Delete column C or 3 for state as it is not needed
    ws.delete_cols(3)

    #ws.delete_cols(7)

    #ws.delete_cols(7)

    #ws.delete_cols(7)

    #ws.delete_cols(7)

    # Determine the number of columns in the sheet
    num_columns = ws.max_column

    # Delete the last 4 columns
    for _ in range(4):
        ws.delete_cols(num_columns)

        # Decrement the column count
        num_columns -= 1

   # Insert a 1 in column 8 (H) to indicate the product is in schematic
    ws.cell(row=1, column=1, value="YES_NO")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):
        for cell in row:
            cell.value = 1

    ## Remove all TBD Remodel from column I
    for cell in ws['F']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('\'', '')

# Rename Columns as required to meet objective for uploading to Snowflake
    ws.cell(row=1, column=1, value='STORE_Name')
    ws.cell(row=1, column=2, value='STORE_Number')
    ws.cell(row=1, column=3, value='UPC')
    ws.cell(row=1, column=4, value='SKU')
    ws.cell(row=1, column=5, value='PRODUCT_NAME')
    ws.cell(row=1, column=6, value='MANUFACTURER')
    ws.cell(row=1, column=7, value='SEGMENT')
    ws.cell(row=1, column=8, value='YES_NO')
    ws.cell(row=1, column=9, value='ACTIVATION_STATUS')
    ws.cell(row=1, column=10, value='COUNTY')
    ws.cell(row=1, column=11, value='CHAIN_NAME')

    # Insert a new column "Chain_Name" at the beginning and fill it with "WALMART"
    #ws.insert_cols(1)
    ws.cell(row=1, column=11, value="CHAIN_NAME")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=11, max_col=11):
        for cell in row:
            cell.value = "TARGET"
 


    return workbook