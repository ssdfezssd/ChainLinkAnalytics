from concurrent.futures.thread import _worker
import streamlit as st
import pandas as pd
import openpyxl
import re
from openpyxl.styles import Border, Side, PatternFill, Font, NamedStyle
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import datetime






def format_SPROUTS_Schedule(workbook):

    st.write("YAY YOU CALLED ME Sprouts")

    # Select the Reset Dates sheet
    ws = workbook['Spring_2023_Beer_Dates']


    # Remove filter from the worksheet
    ws.auto_filter.ref = None

    # Expand all rows in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            ws.row_dimensions[cell.row].hidden = False

    # Unfreeze rows in the worksheet
    ws.freeze_panes = None                

    # Delete Row 1 in the spreadsheet
    ws.delete_rows(1)

    # Insert new column A hold chain name
    ws.insert_cols(1)
    
    # Insert new column F to hold 
    ws.insert_cols(6)

    # Copy data from column C to column F
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=6).value = ws.cell(row=row, column=3).value

    # Remove data from column C starting from the second row
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=3).value = None


    # Copy data from column E to column G
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=7).value = ws.cell(row=row, column=5).value

    # Remove data from column C starting from the second row
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=5).value = None

    # Copy data from column D to column E
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=5).value = ws.cell(row=row, column=4).value

    # Remove data from column D starting from the second row
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=4).value = None

    # Get the maximum row number in column b
    max_row_b = ws.max_row
        
    # Add "SPROUTS" to each cell in column A
    for row in range(2, max_row_b + 1):
        cell = ws.cell(row=row, column=1)
        cell.value = "SPROUTS"

    # Get the maximum row number in column b
    max_row_b = ws.max_row
        
    # Add "SPROUTS" to each cell in column C
    for row in range(2, max_row_b + 1):
        cell = ws.cell(row=row, column=3)
        cell.value = "SPROUTS"

    # Remove data from column I starting from the second row
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=9).value = None

    # Copy data from column M to column I starting from the second row
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=13).value
        ws.cell(row=row, column=9).value = value

     # Remove data from column D starting from the second row
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=13).value = None

    # Delete the current column F
    ws.delete_cols(10)

    # Specify the column you want to format (e.g., column A)
    column_letter = 'J'

    # Iterate over the cells in the column and set the number format
    for cell in ws[column_letter]:
        cell.number_format = 'mm/dd/yyyy'


     # Move cells with "To Follow" from column K to column N
    for row in range(2, ws.max_row + 1):
        cell_k = ws['K{}'.format(row)]
        cell_n = ws['M{}'.format(row)]

        if cell_k.value:
            value = str(cell_k.value)
            match = re.match(r"(To Follow|To follow)\s*#(\d+)", value)
            if match:
                to_follow_number = match.group(2)
                cell_n.value = "To Follow #" + to_follow_number
                cell_k.value = None


    # Enter "7:00 AM" in cells of column K that don't have a time value
    for row in range(2, ws.max_row + 1):
        cell_k = ws['K{}'.format(row)]

        if cell_k.value is None or not isinstance(cell_k.value, datetime.time):
            cell_k.value = "7:00 AM"
            cell_k.number_format = numbers.FORMAT_TEXT


          # Apply gridlines to all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

            # Remove cell colors
            cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

    # Make the entire sheet have black-colored text with font size 11
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(color="00000000", size=11)

    ## Remove all TBD Remodel from column I
    for cell in ws['F']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('\'', '')

    # Rename Columns as required to meet objective for uploading to Snowflake
    ws.cell(row=1, column=1, value='CHAIN_NAME')
    ws.cell(row=1, column=2, value='STORE_NUMBER')
    ws.cell(row=1, column=3, value='STORE_NAME')
    ws.cell(row=1, column=4, value='PHONE')
    ws.cell(row=1, column=5, value='CITY')
    ws.cell(row=1, column=6, value='ADDRESS')
    ws.cell(row=1, column=7, value='STATE')
    ws.cell(row=1, column=8, value='COUNTY')
    ws.cell(row=1, column=9, value='TEAM_LEAD')
    ws.cell(row=1, column=10, value='RESET_DATE')
    ws.cell(row=1, column=11, value='RESET_TIME')
    ws.cell(row=1, column=12, value='STATUS')
    ws.cell(row=1, column=13, value='NOTES')

    # Return the workbook to calling function
    return workbook
