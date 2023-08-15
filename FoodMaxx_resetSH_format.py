import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, NamedStyle
from openpyxl import Workbook
from datetime import datetime, time
from datetime import datetime
import re


def remove_autofilter(ws):
    # Check if autofilter is applied to the worksheet
    if ws.auto_filter:
        # Remove the autofilter
        ws.auto_filter.ref = None



def delete_sheet_by_name(workbook, sheet_name):
    # Find the sheet by name
    sheet = workbook[sheet_name]
    
    # Remove the sheet from the workbook
    workbook.remove(sheet)


def format_FOODMAXX_schedule(workbook):
    
   # Get the worksheet by name ('Food Maxx')
    ws = workbook['Sheet1']

    
    # Get the original worksheet by name ('Sheet1')
    original_ws = workbook['Sheet1']

    delete_sheet_by_name(workbook, 'Sheet2')

    # Create a new worksheet named 'Sheet2'
    new_ws = workbook.create_sheet(title='Sheet2')

    # Copy data from column A to the last column, starting from row 2
    for row in original_ws.iter_rows(min_row=5, min_col=1):
        values = [cell.value for cell in row]
        new_ws.append(values)

    # Remove the original 'Sheet1' worksheet
    workbook.remove(original_ws)

    ws =new_ws

    # Check if the named style "time_format" already exists
    time_format_exists = False
    for style in workbook._named_styles:
        if style.name == 'time_format':
            time_format_exists = True
            break

    # If the named style does not exist, create it
    if not time_format_exists:
        time_format = openpyxl.styles.NamedStyle(name='time_format', number_format='hh:mm AM/PM')
        workbook.add_named_style(time_format)

    # Create new column C to hold Store Name
    ws.insert_cols(3)

       
    # Move data from Column A to Column C
    ws.move_range("A1:A{}".format(ws.max_row), cols=2)

    #Fill column A with "Save Mart" starting from row 2
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.value = "SAVEMART"
    
    # Create new column D to hold Phone
    ws.insert_cols(4)

    # Create new column E to hold City
    ws.insert_cols(5)

    # Move data from Column H to Column E
    #ws.move_range("H1:H{}".format(ws.max_row), cols=4)

    # Copy data from Column H to Column E
    for row_number in range(1, ws.max_row + 1):
        value = ws.cell(row=row_number, column=8).value
        if value:
            ws.cell(row=row_number, column=5, value=value)


    # Remove data from Column H
    for row_number in range(1, ws.max_row + 1):
        ws.cell(row=row_number, column=8, value="")

    # Set the header in Row 1 as "STATE"
    ws.cell(row=1, column=8, value="STATE")

    #Fill column H with "CA" starting from row 2
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.value = "CA"

    # Set the header in Row 1 as "STATE"
    ws.cell(row=1, column=9, value="COUNTY")

    # Create new column J to hold Team Lead
    ws.insert_cols(10)

    # Set the header in Row 1 as "Team Lead"
    ws.cell(row=1, column=10, value="TEAM_LEAD")

    # Copy data from Column F to Column J
    for row_number in range(1, ws.max_row + 1):
        value = ws.cell(row=row_number, column=6).value
        if value:
            ws.cell(row=row_number, column=10, value=value)

    # delete column F "AB WHOLESALE"
    ws.delete_cols(6)

    # Format column J as date "mm/dd/yyyy"
    date_format = NamedStyle(name='date_format', number_format='mm/dd/yyyy')
    for cell in ws['J']:
        cell.style = date_format


    
    # delete column L "Email Contact"
    ws.delete_cols(12)

    # Create a new column (Column L) to hold the formatted time
    ws.insert_cols(12)

        # Copy data from Column K to L where L is empty and format it as "hh:mm AM/PM"
    for row in ws.iter_rows(min_row=2, min_col=11, max_col=12):
        value_k = row[0].value
        value_l = row[1].value
        if not value_l:
            if value_k and isinstance(value_k, (datetime, time)):
                # If the value in column K is already a datetime object or time object, format it directly
                formatted_time = value_k.strftime("%I:%M %p")
            elif value_k:
                # Try to parse the time using different formats
                formats = ["%I:%M %p", "%I:%M%p", "%H:%M"]
                for fmt in formats:
                    try:
                        time_obj = datetime.strptime(value_k, fmt).time()
                        formatted_time = time_obj.strftime("%I:%M %p")
                        break
                    except ValueError:
                        continue
                else:
                    # If none of the formats match, set formatted_time to an empty string
                    formatted_time = ""
            else:
                # If the value in column K is None, set formatted_time to an empty string
                formatted_time = ""

            row[1].value = formatted_time

            # delete column L "Email Contact"
    ws.delete_cols(11)

    # Loop through all cells in Column C
    for cell in ws['C']:
        # Check if the cell has a value (not empty)
        if cell.value:
            # Convert the value to uppercase and remove leading/trailing spaces
            cell.value = cell.value.strip().upper()

        # Insert "01/01/1900" in Column J if it's empty
    for cell in ws['J']:
        if not cell.value:
            cell.value = datetime(1900, 1, 1).strftime("%m/%d/%Y")

    # Insert "06:00 AM" in Column K if it's empty
    for cell in ws['K']:
        if not cell.value:
            cell.value = datetime.strptime("06:00 AM", "%I:%M %p").strftime("%I:%M %p")

# Rename Columns as required to meet objective for uploading to Snowflake
    ws.cell(row=1, column=1, value='CHAIN_NAME')
    ws.cell(row=1, column=2, value='STORE_NUMBER')
    ws.cell(row=1, column=3, value='STORE_NAME')
    ws.cell(row=1, column=4, value='PHONE')
    ws.cell(row=1, column=5, value='CITY')
    ws.cell(row=1, column=6, value='ADDRESS')
    ws.cell(row=1, column=7, value='STATE')
    ws.cell(row=1, column=8, value='COUNTY')
    ws.cell(row=1, column=9, value='TEAM_LEAD')
    ws.cell(row=1, column=10, value='RESET_DATE')
    ws.cell(row=1, column=11, value='RESET_TIME')
    ws.cell(row=1, column=12, value='STATUS')
    ws.cell(row=1, column=13, value='NOTES')

    # Return the path to the saved file
    return workbook








