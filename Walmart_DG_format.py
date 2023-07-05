import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows



import openpyxl

def format_WALMART_DistroGrid(workbook):
    st.write("YAY YOU CALLED ME Walmart dg CODE")

    # Select the Walmart sheet
    ws = workbook['Walmart']

    # Insert a new column "STORE_NAME" at the beginning and fill it with "WALMART"
    ws.insert_cols(1)
    ws.cell(row=1, column=1, value="STORE_NAME")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.value = "Walmart"

    ws.delete_cols(2, 5)

    # Convert the 'Store Number' column from text to numeric format
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value is not None:
                cell.value = int(cell.value)


      ## Remove all TBD Remodel from column I
    for cell in ws['H']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('Add', '1')

          ## Remove all TBD Remodel from column I
    for cell in ws['H']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('Keep', '1')


        rows_to_delete = []

    # Iterate over the rows in column G and identify rows to delete
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        cell = row[0]
        if cell.value == "Delete":
            rows_to_delete.append(cell.row)

    # Delete the identified rows in reverse order to avoid shifting issues
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)

        ws.delete_cols(8)

        # Rename Columns as required to meet objective for uploading to Snowflake
    ws.cell(row=1, column=1, value='STORE_Name')
    ws.cell(row=1, column=2, value='STORE_Number')
    ws.cell(row=1, column=3, value='UPC')
    ws.cell(row=1, column=4, value='SKU')
    ws.cell(row=1, column=5, value='PRODUCT_NAME')
    ws.cell(row=1, column=6, value='MANUFACTURER')
    ws.cell(row=1, column=7, value='SEGMENT')
    ws.cell(row=1, column=8, value='YES_NO')
    ws.cell(row=1, column=9, value='ACTIVATION_STATUS')
    ws.cell(row=1, column=10, value='COUNTY')
    ws.cell(row=1, column=11, value='CHAIN_NAME')

    # Insert a new column "Chain_Name" at the beginning and fill it with "WALMART"
    #ws.insert_cols(1)
    ws.cell(row=1, column=11, value="CHAIN_NAME")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=11, max_col=11):
        for cell in row:
            cell.value = "WALMART"

    # Convert the 'UPC' column from text to numeric format
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            if cell.value is not None:
                cell.value = int(cell.value)


    # Assuming the sheet name is 'Walmart', you can modify it as per your actual sheet name
    sheet = workbook['Walmart']

    # Iterate through the rows starting from the second row
    for row in sheet.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            # Get the value from Column G
            g_value = cell.value

            # Set the value to None in Column G
            cell.value = None

            # Get the corresponding cell in Column H
            h_cell = sheet.cell(row=cell.row, column=8)

            # Set the value from Column G to Column H
            h_cell.value = g_value


    # Assuming the sheet name is 'Walmart', you can modify it as per your actual sheet name
    sheet = workbook['Walmart']

    # Iterate through the rows starting from the second row
    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            # Get the value from Column D
            d_value = cell.value

            # Set the value to None in Column D
            cell.value = None

            # Get the corresponding cell in Column F
            f_cell = sheet.cell(row=cell.row, column=6)

            # Set the value from Column F to Column D
            f_cell.value = d_value


    return workbook
