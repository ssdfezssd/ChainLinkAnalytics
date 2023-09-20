import streamlit as st
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers, PatternFill, Border, Side, Font, NamedStyle
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import openpyxl.utils.datetime as xl_datetime
import datetime
from openpyxl import Workbook
from openpyxl.styles import NamedStyle


#==================================================================================================================
#
#Function to transform Raleys excel file
#
##===============================================================================================================


# Helper function to convert time from text to datetime object
def text_to_time(text):
    try:
        # Convert the cell value to a string
        text = str(text)
        time_obj = datetime.datetime.strptime(text, '%I:%M %p')
        return time_obj.strftime('%H:%M %p')  # Convert back to the desired string format
    except ValueError:
        try:
            # Try converting numeric time fraction to time
            time_fraction = float(text)
            hours = int(time_fraction * 24)
            minutes = int((time_fraction * 24 * 60) % 60)
            time_obj = datetime.time(hours, minutes)
            return time_obj.strftime('%H:%M %p')  # Convert to the desired string format
        except ValueError:
            return text




def format_RALEYS_Schedule(workbook):
    import datetime
    st.write("Formatting workbook for Raleys")
    # Delete all sheets except Reset Dates
    for sheet_name in workbook.sheetnames:
        if sheet_name != 'Reset Dates':
            workbook.remove(workbook[sheet_name])

    # Select the Reset Dates sheet
    ws = workbook['Reset Dates']
    #remove_hidden_rows_and_columns(ws)
    
    # Rename the sheet to "Reset_Dates"
    ws.title = 'Reset_Dates'
    
    # Remove filter from the worksheet
    ws.auto_filter.ref = None
    
    

    # Delete row 2-5
    ws.delete_rows(1,5)

    # Create new column A to hold store number
    ws.insert_cols(1)

     # Create new column B to hold store number
    ws.insert_cols(2)

    # Create new column C to hold store number
    ws.insert_cols(3)

     # Create new column C to hold store number
    ws.insert_cols(8)

    # Create new column H to hold store number
    ws.insert_cols(9)

    # Create new column I to hold store number
    ws.insert_cols(10)

    

     # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Copy data from column F to B
       if row[6] is not None:
            ws.cell(row=index, column=2, value=row[5])

    # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Copy data from column E to C
       if row[4] is not None:
            ws.cell(row=index, column=3, value=row[4])

    # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Copy data from column E to C
       if row[3] is not None:
            ws.cell(row=index, column=8, value=row[3])

    # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Copy data from column G to D
        ws.cell(row=index, column=4, value=row[6])

    # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Copy data from column K to E
        ws.cell(row=index, column=5, value=row[10])

    # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Copy data from column L to F
        ws.cell(row=index, column=6, value=row[11])

    # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, min_col=7), start=2):
    # Clear the cell value in column G
        ws.cell(row=index, column=7).value = None

    # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Copy data from column O to I
        ws.cell(row=index, column=9, value=row[14])

    # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, min_col=7), start=2):
        # Clear the cell value in column K
        ws.cell(row=index, column=11).value = None

   # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, min_col=7), start=2):
        # Clear the cell value in column L
        ws.cell(row=index, column=12).value = None
    
        
    # Remove all TBD Remodel from column J
    for cell in ws['J']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('TBD REMODEL', '1/1/1900')
            
     # Remove all TBD Remodel from column J
    for cell in ws['J']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('The store receives schematics ', '1/1/1900')

    # Remove all - from column J
    for cell in ws['J']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('-', '1/1/1900')

    # Remove all - from column I
    for cell in ws['K']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('-', '6:00 AM')
            
    # Remove all - from column I
    for cell in ws['K']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('from Raleys', '6:00 AM')
            
    # Remove all - from column I
    for cell in ws['K']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('change', '6:00 AM')

    

        

  
    # if ws.dimensions is not None:
    #     # Iterate over column I to convert text to number and format as date
    #     for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
    #         for cell in row:
    #             if cell.value and cell.value.strip():
    #                 try:
    #                     cell.value = float(cell.value)
    #                     cell.number_format = 'mm/dd/yyyy'
    #                     print(f"Converting cell {cell.coordinate} to float: {cell.value}")
    #                     if cell.value == 1:
    #                         cell.value = datetime.datetime(1900, 1, 1)
    #                     else:
    #                         cell.value = datetime.datetime(1900, 1, 1) + datetime.timedelta(days=(cell.value - 1))
    #                     print(f"Resulting value for cell {cell.coordinate}: {cell.value}")
    #                 except ValueError:
    #                     print(f"Failed to convert cell {cell.coordinate} to float.")

       

    # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Copy data from column P to J
        ws.cell(row=index, column=10, value=row[15])    


    from openpyxl.styles import NamedStyle

    ## insert time in empty cells column J
    ## Define the time value to insert
    #time_value = "07:00 AM"

    ## Iterate over the cells in column J
    #for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):
    #    for cell in row:
    #        # Check if the cell is empty
    #        if cell.value is "":
    #            # Assign the time value to the cell
    #            cell.value = time_value


   

 


    ## Create a named style for the time format
    #time_format = NamedStyle(name='time_format2')
    #time_format.number_format = 'hh:mm AM/PM'

    ## Apply the time format to the cells
    #for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):
    #    for cell in row:
    #        if cell.value:
    #            cell.style = time_format


# Remove fill for all columns in row 1
    for cell in ws[1]:
        cell.fill = PatternFill(fill_type=None)

    # Define the border style
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Add grid lines to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

      # Create new column H to hold store number
    ws.insert_cols(4)

    
          # Create new column H to hold store number
    ws.insert_cols(5)

    # Iterate over the rows in column G and copy the values to column E
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            # Get the value from column G
            value = cell.value

            # Calculate the destination column index (E = 5)
            dest_col = 5

            # Get the destination cell in column E
            dest_cell = ws.cell(row=cell.row, column=dest_col)

            # Set the value in the destination cell
            dest_cell.value = value

            # Clear the original cell in column G
            cell.value = None


    # Iterate over the rows in column H and copy the values to column G
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            # Get the value from column G
            value = cell.value

            # Calculate the destination column index (G = 7)
            dest_col = 7

            # Get the destination cell in column G
            dest_cell = ws.cell(row=cell.row, column=dest_col)

            # Set the value in the destination cell
            dest_cell.value = value

            # Clear the original cell in column H
            cell.value = None

    # Iterate over the rows in column J and copy the values to column I
    for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):
        for cell in row:
            # Get the value from column J
            value = cell.value

            # Calculate the destination column index (I = 9)
            dest_col = 9

            # Get the destination cell in column I
            dest_cell = ws.cell(row=cell.row, column=dest_col)

            # Set the value in the destination cell
            dest_cell.value = value

            # Clear the original cell in column J
            cell.value = None

    # Iterate over the rows in column K and copy the values to column J
    for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            # Get the value from column K
            value = cell.value

            # Calculate the destination column index (J = 10)
            dest_col = 10

            # Get the destination cell in column J
            dest_cell = ws.cell(row=cell.row, column=dest_col)

            # Set the value in the destination cell
            dest_cell.value = value

            # Clear the original cell in column K
            cell.value = None

    # Iterate over the rows in column L and copy the values to column K
    for row in ws.iter_rows(min_row=2, min_col=12, max_col=12):
        for cell in row:
            # Get the value from column L
            value = cell.value

            # Calculate the destination column index (K = 11)
            dest_col = 11

            # Get the destination cell in column K
            dest_cell = ws.cell(row=cell.row, column=dest_col)

            # Set the value in the destination cell
            dest_cell.value = value

            # Clear the original cell in column L
            cell.value = None


    # Iterate over the rows in column T and copy the values to column M
    for row in ws.iter_rows(min_row=2, min_col=20, max_col=20):
        for cell in row:
            # Get the value from column T
            value = cell.value

            # Calculate the destination column index (M = 13)
            dest_col = 13

            # Get the destination cell in column M
            dest_cell = ws.cell(row=cell.row, column=dest_col)

            # Set the value in the destination cell
            dest_cell.value = value

            # Clear the original cell in column T
            cell.value = None

    # Create a named style for the time format
    time_format = NamedStyle(name='time_format')
    time_format.number_format = 'hh:mm AM/PM'

    # Apply the time format to the cells
    for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            if cell.value:
                # Convert the value to the desired time format string
                cell.value = text_to_time(cell.value)
                cell.style = time_format

        
            
            # Apply the time format to the cells
    for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            if cell.value:
                cell.value = text_to_time(cell.value)

    # # Create a named style for the time format
    #time_format = NamedStyle(name='time_format3')
    #time_format.number_format = 'hh:mm AM/PM'

    ## Apply the time format to the cells
    #for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
    #    for cell in row:
    #        if cell.value:
    #            cell.style = time_format

    

    # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        # Remove column N
        ws.delete_cols(openpyxl.utils.column_index_from_string('N'))

        # Remove column O
        ws.delete_cols(openpyxl.utils.column_index_from_string('O'))

        # Remove column P
        ws.delete_cols(openpyxl.utils.column_index_from_string('P'))

        # Remove column Q
        ws.delete_cols(openpyxl.utils.column_index_from_string('Q'))

        # Remove column R
        ws.delete_cols(openpyxl.utils.column_index_from_string('R'))

        # Remove column S
        ws.delete_cols(openpyxl.utils.column_index_from_string('S'))

        # Remove column T
        ws.delete_cols(openpyxl.utils.column_index_from_string('T'))


        # insert time in empty cells column J
    # Define the time value to insert
    time_value = "1900-01-01"

    # Iterate over the cells in column J
    for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            # Check if the cell is empty
            if cell.value is "":
                # Assign the time value to the cell
                cell.value = time_value

# Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        # Specify the column you want to format (e.g., column A)
        column_letter = 'J'

    # Iterate over the cells in the column and set the number format
    for cell in ws[column_letter]:
        cell.number_format = 'mm/dd/yyyy'

    # Get the maximum row number in column A
    max_row = ws.max_row

    # Add "RALEYS" to each cell in column A
    for row in range(1, max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.value = "RALEYS"


     # Remove all Raleys Supermarket and repalce with RALEYS from column C
    for cell in ws['C']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('Raleys Supermarket', 'RALEYS')

    # Remove all Raleys Supermarket and repalce with RALEYS from column C
    for cell in ws['C']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('Bel Air Market', 'Bel Air')



    #  Remove all Raleys Supermarket and repalce with RALEYS from column C
    for cell in ws['C']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('Nob Hill Foods', 'NOB HILL')

        #  Remove all Raleys Supermarket and repalce with RALEYS from column C
    for cell in ws['I']:
        if cell.value is not None:
            cell.value = str(cell.value).replace(',', ' ')


    # Iterate over all columns in the worksheet
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

    # Iterate over each cell in the column and calculate the maximum length
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except TypeError:
            pass

    # Adjust the column width based on the maximum length
    adjusted_width = (max_length + 2) * 1.2  # Add some padding and scale factor
    ws.column_dimensions[column_letter].width = adjusted_width

    # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

    
        # Specify the column to format (e.g., column J)
        column_letter = 'J'
    column_index = openpyxl.utils.column_index_from_string(column_letter)

    from datetime import datetime

    # Iterate over the cells in the column and convert them to date format
    for row in ws.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
        for cell in row:
            if isinstance(cell.value, datetime):  # Check if the cell value is already a datetime object
                cell.number_format = 'mm/dd/yyyy'  # Set the number format
            else:
                # Convert the cell value to a datetime object
                try:
                    date_value = datetime.strptime(str(cell.value)[:8], '%Y%m%d')
                    cell.value = date_value
                    cell.number_format = 'mm/dd/yyyy'  # Set the number format
                except ValueError:
                    # Handle any invalid or non-date values
                    pass

    
    for cell in ws['K']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('TBD', '06:00 AM')

    # Remove all - from column K
    for cell in ws['K']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('-', '06:00 AM')

    ## Save the modified workbook
    #workbook.save("updated_workbook.xlsx")

    # Iterate through the rows starting from row 2
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        # Rename Columns as required to meet objective for uploading to Snowflake
        ws.cell(row=1, column=1, value='CHAIN_NAME')
        ws.cell(row=1, column=2, value='STORE_NUMBER')
        ws.cell(row=1, column=3, value='STORE_NAME')
        ws.cell(row=1, column=4, value='PHONE')
        ws.cell(row=1, column=5, value='CITY')
        ws.cell(row=1, column=6, value='ADDRESS')
        ws.cell(row=1, column=7, value='STATE')
        ws.cell(row=1, column=8, value='COUNTY')
        ws.cell(row=1, column=9, value='TEAM_LEAD')
        ws.cell(row=1, column=10, value='RESET_DATE')
        ws.cell(row=1, column=11, value='RESET_TIME')
        ws.cell(row=1, column=12, value='STATUS')
        ws.cell(row=1, column=13, value='NOTES')






       
    return workbook


#===================================================================================================================
# End Reformat Excel spreadsheet for Raleys

#====================================================================================================================