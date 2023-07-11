from ctypes import wstring_at
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, NamedStyle, numbers    
from openpyxl.utils import get_column_letter
import datetime


#=======================================================================================================================
# Function to adjust the width of the columns in the sheet
#======================================================================================================================

def adjust_column_widths(worksheet):
    # Iterate over all columns
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        # Find the maximum length of data in each column
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass

        # Set the column width based on the maximum length
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

#=======================================================================================================================
# End of Function to adjust the width of the columns in the sheet
#======================================================================================================================


# Remove cell fills
def remove_cell_fills(worksheet, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Format row heights
def format_row_heights(worksheet, start_row, end_row, height):
    for row in range(start_row, end_row + 1):
        worksheet.row_dimensions[row].height = height




#===================================================================================================================
# Function to reformat Excel spreadsheet for SMART & FINAL
#====================================================================================================================
def format_SMARTFINAL_Schedule(workbook):

    st.write("YAY YOU CALLED ME SMART-N-FINAL")

    # Select the Reset Dates sheet
    ws = workbook['Reset Calendar']

    # Remove filter from the worksheet
    ws.auto_filter.ref = None

    # Remove frozen panes
    ws.freeze_panes = None

    # Create a list to store merged cell ranges
    merged_cell_ranges = []

    # Collect merged cell ranges
    for merged_cell_range in ws.merged_cells.ranges:
        merged_cell_ranges.append(merged_cell_range)

    # Unmerge cells
    for merged_cell_range in merged_cell_ranges:
        ws.unmerge_cells(str(merged_cell_range))

    # Iterate through all images in the sheet and delete them
    for image in ws._images:
        ws._images.remove(image)

    # Delete the first row in the sheet
    ws.delete_rows(1,4)

    # Determine the number of rows with data
    sheet_name = "Reset Calendar"  # Update with the actual sheet name
    ws = workbook[sheet_name]
    max_row = ws.max_row
    for row in reversed(range(1, max_row + 1)):
        if ws.cell(row=row, column=3).value:
            max_row = row
            break

   # Delete extra rows beyond the data range
    if ws.max_row > max_row:
        ws.delete_rows(max_row + 1, ws.max_row - max_row)

    # Insert New column A to hold 'CHAIN_NAME'
    ws.insert_cols(1)

   # Delete all data in column C
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.value = None

    # Insert 'STORE_NAME' in cell A1
    ws['A1'] = 'CHAIN_NAME'

    # Fill the rest of the cells in column C with 'SMART & FINAL'
    max_row = ws.max_row
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, max_row=max_row):
        for cell in row:
            cell.value = 'SMART & FINAL'

    # Insert 'STORE_NAME' in cell C1
    ws['C1'] = 'STORE_NAME'

    # Fill the rest of the cells in column C with 'SMART & FINAL'
    max_row = ws.max_row
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, max_row=max_row):
        for cell in row:
            cell.value = 'SMART & FINAL'

    # Insert New column D to hold 'PHONE_NUMBER'
    ws.insert_cols(4)

    # Insert 'PHONE_NUMBER' in cell D1
    ws['D1'] = 'PHONE_NUMBER'

    
     # Iterate through the rows and swap the values between columns E and F
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=5, max_col=6):
        cell1 = row[0]
        cell2 = row[1]
        cell1.value, cell2.value = cell2.value, cell1.value


    # Insert 'COUNTY' in cell H1
    ws['H1'] = 'COUNTY'

    # Delete all data in column H
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.value = None

    # Insert New column A to hold 'CHAIN_NAME'
    ws.insert_cols(9)

    # Insert 'TEAM_LEAD' in cell I1
    ws['I1'] = 'TEAM_LEAD'

    # Insert 'TEAM_LEAD' in cell L1
    ws['L1'] = 'STATUS'

    # Insert 'TEAM_LEAD' in cell M1
    ws['M1'] = 'NOTES'
    
    # Example usage
    ws= workbook.active  # Assuming you have the worksheet object
    start_row = 1
    end_row = 10
    start_col = 1
    end_col = 5

    # Remove cell fills
    remove_cell_fills(ws, start_row, end_row, start_col, end_col)

        # Format row heights
    format_row_heights(ws, start_row, end_row, 20)
   

    # Rename Columns as required to meet objective for uploading to Snowflake
    ws.cell(row=1, column=1, value='CHAIN_NAME')
    ws.cell(row=1, column=2, value='STORE_NUMBER')
    ws.cell(row=1, column=3, value='STORE_NAME')
    ws.cell(row=1, column=4, value='PHONE')
    ws.cell(row=1, column=5, value='CITY')
    ws.cell(row=1, column=6, value='ADDRESS')
    ws.cell(row=1, column=7, value='STATE')
    ws.cell(row=1, column=8, value='COUNTY')
    ws.cell(row=1, column=9, value='TEAM_LEAD')
    ws.cell(row=1, column=10, value='RESET_DATE')
    ws.cell(row=1, column=11, value='RESET_TIME')
    ws.cell(row=1, column=12, value='STATUS')
    ws.cell(row=1, column=13, value='NOTES')
    
    # Save the modified workbook
    workbook.save('your_modified_workbook.xlsx')


    return workbook