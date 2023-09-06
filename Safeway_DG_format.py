import streamlit as st
import pandas as pd
import openpyxl





# Reformat Excel spreadsheet for SAFEWAY
#====================================================================================================================
def format_SAFEWAY_DistroGrid(workbook):

    st.write("YAY YOU CALLED ME safeway dg CODE")


    # Select the Reset Dates sheet
    ws = workbook['Safeway NorCal Fall Assortment']


    # Get the maximum row number in column B (STORE_Number)
    max_row_b = ws.max_row

    # Find the last row with data in column B (STORE_Number)
    last_row_with_data = max_row_b
    for row in range(max_row_b, 1, -1):
        if ws.cell(row=row, column=2).value is not None:
            last_row_with_data = row
            break

    # Remove rows below the dataset
    for row in range(last_row_with_data + 1, ws.max_row + 1):
        ws.delete_rows(row)

    # Update the maximum row number after deletion
    max_row_b = ws.max_row

    # Insert a new column "STORE_NAME" at the beginning and fill it with "SAFEWAY"
    ws.insert_cols(1)
    ws.cell(row=1, column=1, value="STORE_NAME")
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
    cell.value = "SAFEWAY"

    # Delete columns 3(C)
    ws.delete_cols(3)

    # Delete columns 3(C)
    ws.delete_cols(3)

    # Delete columns 3(C)
    ws.delete_cols(3)

        # Iterate over the rows starting from row 2 and remove all values 
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.value = None

       # Iterate over the rows starting from row 2 and remove all values 
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            cell.value = None

       # Iterate over the rows starting from row 2 and remove all values 
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.value = None

       # Iterate over the rows starting from row 2 and remove all values 
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
        for cell in row:
            cell.value = None


    # Remove the filter from the sheet
    ws.auto_filter.ref = None

    ## Change ADDED and MAINTAIN to 1 IN Column H
    for cell in ws['H']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('ADDED', '1').replace('MAINTAIN', '1')

     # Get the maximum row number in column b
    max_row_b = ws.max_row

    # Add "SAFEWAY" to each cell in column A
    for row in range(2, max_row_b + 1):
        cell = ws.cell(row=row, column=11)
        cell.value = "SAFEWAY"

    # Add "SAFEWAY" to each cell in column A
    for row in range(2, max_row_b + 1):
        cell = ws.cell(row=row, column=1)
        cell.value = "SAFEWAY"

    



    # Rename Columns as required to meet objective for uploading to Snowflake
    ws.cell(row=1, column=1, value='STORE_NAME')
    ws.cell(row=1, column=2, value='STORE_NUMBER')
    ws.cell(row=1, column=3, value='UPC')
    ws.cell(row=1, column=4, value='SKU')
    ws.cell(row=1, column=5, value='PRODUCT_NAME')
    ws.cell(row=1, column=6, value='MANUFACTURER')
    ws.cell(row=1, column=7, value='SEGMENT')
    ws.cell(row=1, column=8, value='YES_NO')
    ws.cell(row=1, column=9, value='ACTIVATION_STATUS')
    ws.cell(row=1, column=10, value='COUNTY')
    ws.cell(row=1, column=11, value='CHAIN_NAME')


    return workbook