import streamlit as st
import pandas as pd
from openpyxl.styles import Border, Side, PatternFill, Font, NamedStyle




def format_Savemart_Schedule(workbook):

    st.write("YAY YOU CALLED ME Save Mart")

    # Select the Reset Dates sheet
    ws = workbook['Save_Mart']

     # Remove filter from the worksheet
    ws.auto_filter.ref = None

    # Delete row 1, 2, 3 
    ws.delete_rows(1, 3)

    # Insert new column A hold chain name
    ws.insert_cols(1)

    # Get the maximum row number in column b
    max_row_b = ws.max_row
        
    # Add "SPROUTS" to each cell in column A
    for row in range(2, max_row_b + 1):
        cell = ws.cell(row=row, column=1)
        cell.value = "SAVEMART"

    # Remove data from column D starting from the second row
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=8).value = None

        # Copy data from column C to column H starting from the second row
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=3).value
        ws.cell(row=row, column=8).value = value

    # Remove data from column D starting from the second row
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=3).value = None

    # Get the maximum row number in column b
    max_row_b = ws.max_row

    # Add "SPROUTS" to each cell in column A
    for row in range(2, max_row_b + 1):
        cell = ws.cell(row=row, column=3)
        cell.value = "SAVEMART"

    # Delete the current column D
    ws.delete_cols(4)  
    
    # Insert new column A hold chain name
    ws.insert_cols(5)

    # Copy data from column G to column E starting from the second row
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=7).value # copy data from column
        ws.cell(row=row, column=5).value = value # copy data to column


    # Remove data from column D starting from the second row
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=7).value = None


    # Insert new column A hold chain name
    ws.insert_cols(8)

    
    # Get the maximum row number in column F
    max_row_f = ws.max_row


    # Add "SPROUTS" to each cell in column A
    for row in range(2, max_row_f + 1):
        cell = ws.cell(row=row, column=7)
        cell.value = "CA"


    # Delete columns N through X
    ws.delete_cols(14, 23)

    # Apply gridlines to all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

            # Remove cell colors
            cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

    # Make the entire sheet have black-colored text with font size 11
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(color="00000000", size=11)


# Format column J as date "mm/dd/yyyy"
    date_format = NamedStyle(name='date_format')
    date_format.number_format = 'mm/dd/yyyy'
    column_letter = 'J'
    for cell in ws[column_letter]:
        cell.style = date_format

    # Find the last row with data in column A
    last_row = ws.max_row
    for row in reversed(range(1, last_row + 1)):
        if ws.cell(row=row, column=1).value:
            last_row = row
            break

    # Fill in empty cells in column J with '01/01/1900'
    for row in range(1, last_row + 1):
        if not ws.cell(row=row, column=10).value:
            ws.cell(row=row, column=10).value = '01/01/1900'

    # Fill in empty cells in column J with '6:00 AM'
    for row in range(2, last_row + 1):
        if not ws.cell(row=row, column=11).value:
            ws.cell(row=row, column=11).value = '6:00 AM'

    # Delete rows past the last row with data in column A
    if last_row < ws.max_row:
        ws.delete_rows(last_row + 1, ws.max_row - last_row)


    # Rename Columns as required to meet objective for uploading to Snowflake
    ws.cell(row=1, column=1, value='CHAIN_NAME')
    ws.cell(row=1, column=2, value='STORE_NUMBER')
    ws.cell(row=1, column=3, value='STORE_NAME')
    ws.cell(row=1, column=4, value='PHONE')
    ws.cell(row=1, column=5, value='CITY')
    ws.cell(row=1, column=6, value='ADDRESS')
    ws.cell(row=1, column=7, value='STATE')
    ws.cell(row=1, column=8, value='COUNTY')
    ws.cell(row=1, column=9, value='TEAM_LEAD')
    ws.cell(row=1, column=10, value='RESET_DATE')
    ws.cell(row=1, column=11, value='RESET_TIME')
    ws.cell(row=1, column=12, value='STATUS')
    ws.cell(row=1, column=13, value='NOTES')



    # Return the path to the saved file
    return workbook

